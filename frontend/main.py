from fastapi import FastAPI, UploadFile, Request, Form, Body, Path, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import uvicorn
import os
import shutil
import subprocess
import pandas as pd
import csv
import sys
import json
from typing import List, Dict, Any, Optional
import io
import math
import re
import numpy as np
import logging
from datetime import datetime
import config # <--- ДОБАВИТЬ, если отсутствует
from pydantic import BaseModel # <--- ДОБАВЛЕН ИМПОРТ
# --- Добавляем импорты для стилизации Excel ---
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side # <<< ДОБАВЛЕНЫ Border, Side
# --- Конец импортов для стилизации ---
# --- Добавляем импорты для исправления XLSX ---
import zipfile
import tempfile

# Добавляем путь к корню проекта
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

# Импортируем напрямую из файла в корне проекта
from geocoder import geocode_address
# --- Импортируем функцию расчета расстояний --- 
import route_distance 
# --- ДОБАВЛЕН ИМПОРТ get_api_key --- 
from utils import ensure_data_dirs, get_api_key 
# --- КОНЕЦ ДОБАВЛЕНИЯ ---

app = FastAPI()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.abspath(os.path.join(BASE_DIR, ".."))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploaded")
DATA_FOLDER = os.path.join(BASE_DIR, "data")
PROJECT_DATA_FOLDER = os.path.join(ROOT_DIR, "data")
GEOCODED_RESULTS_FOLDER = os.path.join(PROJECT_DATA_FOLDER, "geocoded_results")
ROUTE_RESULTS_FOLDER = os.path.join(PROJECT_DATA_FOLDER, "route_results")
PARSED_ADDRESSES_FOLDER = os.path.join(PROJECT_DATA_FOLDER, "parsed_addresses")
SUMMARY_CSV_FOLDER = os.path.join(PROJECT_DATA_FOLDER, "summary_csv")

# Создаем все необходимые директории
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DATA_FOLDER, exist_ok=True)
os.makedirs(GEOCODED_RESULTS_FOLDER, exist_ok=True)
os.makedirs(ROUTE_RESULTS_FOLDER, exist_ok=True)
os.makedirs(PARSED_ADDRESSES_FOLDER, exist_ok=True)
os.makedirs(SUMMARY_CSV_FOLDER, exist_ok=True)

app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

# Вспомогательная функция для создания безопасных имен файлов
def sanitize_filename(filename):
    """Преобразует строку в безопасное имя файла"""
    if not isinstance(filename, str):
        return "unnamed"
    # Заменяем пробелы на подчеркивания, сохраняя все остальные символы
    s = filename.strip().replace(' ', '_')
    return s if s else "unnamed"

# Вспомогательная функция для очистки данных перед JSON сериализацией
def sanitize_data_for_json(data):
    if isinstance(data, dict):
        return {k: sanitize_data_for_json(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_data_for_json(item) for item in data]
    elif isinstance(data, float):
        if math.isnan(data) or math.isinf(data):
            return None # Заменяем NaN/inf на None
        return data
    else:
        return data

# --- Утилита для исправления регистра sharedStrings.xml (из test_excel_read.py) ---
def fix_xlsx_casing(original_path):
    """Создает временную копию XLSX, если sharedStrings.xml имеет неправильный регистр.
    Возвращает путь к исправленной копии или исходный путь. None при ошибке.
    """
    expected = 'xl/sharedStrings.xml'
    incorrect_candidate = 'xl/SharedStrings.xml'
    fixed_path = original_path
    
    try:
        with zipfile.ZipFile(original_path, 'r') as zin:
            archive_files = zin.namelist()
            if expected in archive_files:
                return original_path # Исправление не нужно
            
            if incorrect_candidate not in archive_files:
                 # Возвращаем исходный путь, т.к. проблема не в регистре или файл не xlsx
                 return original_path 
            
            # --- Исправление требуется ---
            print(f"INFO: Исправление регистра {incorrect_candidate}...")
            temp_fd, fixed_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd) # Закрываем дескриптор, нам нужен только путь

            # Используем вложенный with для гарантии закрытия
            with zipfile.ZipFile(fixed_path, 'w', zipfile.ZIP_DEFLATED) as zout, \
                 zipfile.ZipFile(original_path, 'r') as zin_inner:
                for item in zin_inner.infolist():
                    buffer = zin_inner.read(item.filename)
                    # Переименовываем при копировании
                    target_name = expected if item.filename == incorrect_candidate else item.filename
                    # Создаем ZipInfo с правильным именем/исходными атрибутами
                    target_info = zipfile.ZipInfo(target_name, date_time=item.date_time)
                    target_info.compress_type = item.compress_type
                    target_info.external_attr = item.external_attr 
                    target_info.internal_attr = item.internal_attr
                    zout.writestr(target_info, buffer)
            print(f"INFO: Создан временный файл {fixed_path}")
            return fixed_path

    except zipfile.BadZipFile:
        print(f"WARNING: Не удалось прочитать {original_path} как ZIP-архив.")
        # Возвращаем None, чтобы сигнализировать об ошибке чтения как XLSX
        return None 
    except Exception as e:
        print(f"ERROR: Ошибка обработки {original_path} при исправлении регистра: {e}")
        # Удаляем временный файл при ошибке, если он создался
        if fixed_path != original_path and os.path.exists(fixed_path):
             try: os.remove(fixed_path)
             except Exception: pass
        return None # В случае серьезной ошибки возвращаем None
# --- Конец утилиты --- 

# --- ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ФОРМАТИРОВАНИЯ (можно вынести в utils.py если нужно) ---
def format_distance_data_from_segments(segments_data):
    """Форматирует данные о расстоянии и времени из списка сегментов."""
    total_distance_meters = 0
    total_duration_seconds = 0
    formatted_segments = [] # Можно добавить форматирование сегментов, если нужно фронтенду
    
    for segment in segments_data:
        # Используем .get с 0 по умолчанию
        distance = segment.get('distance', 0)
        duration = segment.get('duration', 0)
        
        # Пропускаем сегменты с ошибками (где distance/duration могли остаться 0 из-за ошибки API)
        if segment.get('status') != "OK":
            print(f"  - Warning: Skipping segment {segment.get('from_index', '?')}->{segment.get('to_index', '?')} due to status: {segment.get('status')}")
            continue # Не учитываем ошибочные сегменты в сумме
            
        if isinstance(distance, (int, float)):
            total_distance_meters += distance
        if isinstance(duration, (int, float)):
            total_duration_seconds += duration
        
        # Добавляем сегмент в список (если нужно)
        # formatted_segments.append({...}) 

    # Округляем расстояние до КМ (как в get_route_data_endpoint)
    total_distance_km_rounded = round(total_distance_meters / 1000) if total_distance_meters is not None else 0
    formatted_distance = f"{total_distance_km_rounded} км"

    # Форматируем время
    hours = total_duration_seconds // 3600
    minutes = (total_duration_seconds % 3600) // 60
    formatted_duration = f"{hours} ч {minutes} мин" if hours > 0 else f"{minutes} мин"
    if total_duration_seconds == 0: # Если время 0, показать 0 мин
        formatted_duration = "0 мин" 

    return {
        "segments": segments_data, # Возвращаем исходные сегменты
        "total_distance": total_distance_km_rounded, # Округленные КМ
        "total_duration": total_duration_seconds,  # Секунды
        "formatted_distance": formatted_distance,
        "formatted_duration": formatted_duration
    }
# --- КОНЕЦ ВСПОМОГАТЕЛЬНОЙ ФУНКЦИИ ---

# Структура для хранения обработанных данных
class RouteData:
    def __init__(self):
        self.routes = {}  # Ключ: sanitized_name
        self.summary = {} # Ключ: sanitized_name, Значение: {...}
        self.drivers = {} # <--- НОВОЕ ПОЛЕ: Словарь для хранения водителей {sanitized_name: driver_name}
        self.current_file = ""  # Имя текущего файла
        self.global_service_time_minutes = 0 # <--- НОВОЕ ГЛОБАЛЬНОЕ ПОЛЕ (по умолчанию 0)
        self.report_date_str: Optional[str] = None # <--- ДОБАВЛЕНО ПОЛЕ ДЛЯ ДАТЫ ОТЧЕТА

    def add_route(self, name, data):
        sanitized_name = sanitize_filename(name)
        if not sanitized_name: return

        # Сохраняем полные данные маршрута (geocoder_output, route_points, distance_data)
        self.routes[sanitized_name] = data

        distance_data = data.get("distance_data", {})
        distance_km = distance_data.get("total_distance") # Ожидаем округленное до целых км
        duration_sec = distance_data.get("total_duration") # Ожидаем секунды
        duration_formatted = distance_data.get("formatted_duration", "Н/Д")
        # --- ИЗМЕНЕНО: Используем переданное количество точек, если оно есть ---
        # Старая логика: num_intermediate_stops = len(data.get("route_points", []))
        # Новая логика: 
        # Пытаемся получить 'number_of_stops' из входящих данных (это должно быть кол-во промежуточных точек)
        num_intermediate_stops = data.get("number_of_stops") 
        if num_intermediate_stops is None:
             # Если не передано, используем старую логику как fallback, но она может быть неточной
             # (т.к. route_points может содержать старт/финиш)
             print(f"[add_route] WARNING: 'number_of_stops' not provided in data for route '{name}'. Falling back to len(route_points). This might be incorrect.")
             # Пытаемся угадать: если в route_points есть старт/финиш, вычитаем 2
             # Это НЕнадежно, лучше всегда передавать number_of_stops!
             route_points_list = data.get("route_points", [])
             if len(route_points_list) >= 2 and route_points_list[0] == route_points_list[-1]: # Простой эвристический тест
                 num_intermediate_stops = len(route_points_list) - 2
             else:
                 num_intermediate_stops = len(route_points_list)
        # --- КОНЕЦ ИЗМЕНЕНИЯ ---

        # Инициализация или обновление записи в summary
        if sanitized_name not in self.summary:
            self.summary[sanitized_name] = {
                "original_name": name,
                "distance": distance_km if distance_km is not None else "Н/Д",
                "duration_seconds": duration_sec, # Храним секунды для расчета разницы и времени на маршруте
                "duration_formatted": duration_formatted, # Время в пути
                "number_of_stops": num_intermediate_stops, # <--- НОВОЕ ПОЛЕ: кол-во точек
                "report_distance": None,
                "report_duration_hours": None,
                "report_duration_minutes": None,
                "distance_difference": None,
                "time_difference_formatted": None,
                "total_route_time_seconds": None, # <--- НОВОЕ ПОЛЕ: общее время в сек
                "total_route_time_formatted": None # <--- НОВОЕ ПОЛЕ: общее время формат.
            }
        else:
            # Обновляем расчетные данные, сохраняя введенные пользователем
            self.summary[sanitized_name]["distance"] = distance_km if distance_km is not None else "Н/Д"
            self.summary[sanitized_name]["duration_seconds"] = duration_sec
            self.summary[sanitized_name]["duration_formatted"] = duration_formatted
            self.summary[sanitized_name]["number_of_stops"] = num_intermediate_stops # Обновляем кол-во точек

        # Пересчитываем все поля, включая новое "Время на маршруте"
        self._recalculate_summary_fields(sanitized_name)

    def get_route_names(self):
        return [data.get("original_name", key) for key, data in self.summary.items()]

    def get_route(self, name):
        sanitized_name = sanitize_filename(name)
        return self.routes.get(sanitized_name, {})

    def update_summary_item(self, name, report_distance=None, report_hours=None, report_minutes=None, comment=None):
        """Обновляет данные для одного маршрута в сводке (без времени на точку)."""
        sanitized_name = sanitize_filename(name)
        if not sanitized_name or sanitized_name not in self.summary:
            print(f"⚠️ Попытка обновить несуществующий маршрут в сводке: {name} (sanitized: {sanitized_name})")
            return False

        summary_item = self.summary[sanitized_name]
        updated = False

        # Обновляем расстояние по отчету
        if report_distance is not None:
            try:
                report_distance_float = float(report_distance) if str(report_distance).strip() != "" else None
                if summary_item.get("report_distance") != report_distance_float:
                    summary_item["report_distance"] = report_distance_float
                    updated = True
            except (ValueError, TypeError):
                if summary_item.get("report_distance") is not None:
                     summary_item["report_distance"] = None # Сбрасываем если некорректный ввод
                     updated = True

        # --- Обновляем время по отчету ---
        new_report_h = None
        new_report_m = None
        try:
            if report_hours is not None and str(report_hours).strip() != "":
                new_report_h = int(report_hours)
            if report_minutes is not None and str(report_minutes).strip() != "":
                new_report_m = int(report_minutes)

            if summary_item.get("report_duration_hours") != new_report_h or \
               summary_item.get("report_duration_minutes") != new_report_m:
                summary_item["report_duration_hours"] = new_report_h
                summary_item["report_duration_minutes"] = new_report_m
                updated = True
                print(f"   Updated report time in summary_item: h={new_report_h}, m={new_report_m}")

        except (ValueError, TypeError):
            if summary_item.get("report_duration_hours") is not None or \
               summary_item.get("report_duration_minutes") is not None:
                summary_item["report_duration_hours"] = None
                summary_item["report_duration_minutes"] = None
                updated = True
                print(f"   Reset report time in summary_item due to conversion error")
                
        # Обновляем комментарий
        if comment is not None and summary_item.get("comment") != comment:
            summary_item["comment"] = comment
            updated = True

        # Пересчитываем поля (включая время на маршруте) и сохраняем, если что-то обновилось
        if updated:
            self._recalculate_summary_fields(sanitized_name) # Пересчет включает новое поле
            self.save_to_disk()
            return True

        return False # Ничего не изменилось

    def _recalculate_summary_fields(self, sanitized_name, summary_dict=None):
        """Пересчитывает разницу в расстоянии, времени и общее время на маршруте."""
        target_summary = summary_dict if summary_dict is not None else self.summary

        if sanitized_name not in target_summary:
            return

        item = target_summary[sanitized_name]
        original_name = item.get("original_name", sanitized_name)
        print(f"-- Recalculating fields for: {original_name}")

        # --- Разница расстояний ---
        report_dist = item.get("report_distance")
        route_dist = item.get("distance")
        if report_dist is not None and isinstance(route_dist, (int, float)):
            try:
                item["distance_difference"] = round(route_dist - float(report_dist), 2)
            except (ValueError, TypeError):
                item["distance_difference"] = None
        else:
            item["distance_difference"] = None
        # print(f"   Distance diff calculated: {item['distance_difference']}") # Убрал лог для краткости

        # --- Расчет Общего Времени на Маршруте (ПРАВИЛЬНОЕ МЕСТО) ---
        total_route_time_seconds = None # Инициализируем переменную
        route_sec = item.get("duration_seconds") # Время в пути
        service_time_sec = 0
        if self.global_service_time_minutes is not None:
            try:
                # Используем .get с дефолтом 0 для number_of_stops
                # --- DEBUG ЛОГ: Какое количество точек используется --- 
                print(f"--- [DEBUG _recalculate_summary_fields] Route: {original_name}, Calculating service time using number_of_stops = {item.get('number_of_stops', 'N/A')}")
                # --- КОНЕЦ DEBUG ЛОГА ---
                service_time_sec = int(self.global_service_time_minutes) * 60 * item.get("number_of_stops", 0)
            except (ValueError, TypeError):
                 service_time_sec = 0

        if route_sec is not None:
            try:
                total_route_time_seconds = int(route_sec) + service_time_sec # <-- Расчет общего времени
                item["total_route_time_seconds"] = total_route_time_seconds # Сохраняем в item
                hours = total_route_time_seconds // 3600
                minutes = (total_route_time_seconds % 3600) // 60
                item["total_route_time_formatted"] = f"{hours} ч {minutes} мин" # Сохраняем формат
            except (ValueError, TypeError) as e:
                print(f"   Error calculating total route time for {original_name}: {e}")
                item["total_route_time_seconds"] = None
                item["total_route_time_formatted"] = None
        else:
            item["total_route_time_seconds"] = None
            item["total_route_time_formatted"] = None
            # print(f"   Skipping total route time calculation (route_sec={route_sec})") # Можно вернуть для отладки

        # --- Разница времени (отчет vs ОБЩЕЕ ВРЕМЯ НА МАРШРУТЕ) ---
        report_h = item.get("report_duration_hours")
        report_m = item.get("report_duration_minutes")
        # route_sec БОЛЬШЕ НЕ НУЖЕН ЗДЕСЬ, используем total_route_time_seconds

        # ... (логика расчета time_difference_formatted остается прежней) ...
        try:
            report_h_int = int(report_h) if report_h is not None and str(report_h).strip() != "" else 0
            report_m_int = int(report_m) if report_m is not None and str(report_m).strip() != "" else 0
            is_report_time_valid = not (report_h_int == 0 and report_m_int == 0 and (report_h is None or str(report_h).strip() == "") and (report_m is None or str(report_m).strip() == ""))
        except (ValueError, TypeError):
             is_report_time_valid = False

        calculate_time_diff = False
        # Проверяем, что ОБЩЕЕ время рассчитано
        if total_route_time_seconds is not None:
            if is_report_time_valid:
                 calculate_time_diff = True
            elif report_h_int == 0 and report_m_int == 0: # Если время отчета не введено (0ч 0м), тоже считаем разницу
                 calculate_time_diff = True # <-- ИЗМЕНЕНО: Считаем разницу, даже если время отчета 0, чтобы показать полное время маршрута как разницу

        if calculate_time_diff:
            try:
                report_total_seconds = report_h_int * 3600 + report_m_int * 60
                # ИСПОЛЬЗУЕМ ОБЩЕЕ ВРЕМЯ НА МАРШРУТЕ для расчета разницы
                time_diff_value_seconds = total_route_time_seconds - report_total_seconds # <-- РЕАЛЬНОЕ ЗНАЧЕНИЕ РАЗНИЦЫ
                sign = "+" if time_diff_value_seconds >= 0 else "-"
                diff_seconds_abs = abs(time_diff_value_seconds) # Используем abs() для форматирования
                diff_hours = diff_seconds_abs // 3600
                diff_minutes = (diff_seconds_abs % 3600) // 60
                item["time_difference_formatted"] = f"{sign}{diff_hours} ч {diff_minutes} мин"
                item["time_difference_seconds"] = time_diff_value_seconds # <-- СОХРАНЯЕМ ЗНАЧЕНИЕ В СЕКУНДАХ
            except (ValueError, TypeError) as e:
                print(f"   Error calculating time diff for {original_name}: {e}")
                item["time_difference_formatted"] = None
                item["time_difference_seconds"] = None # <-- Сбрасываем и секунды
        else:
            item["time_difference_formatted"] = None
            item["time_difference_seconds"] = None # <-- Сбрасываем и секунды

    def get_summary(self):
        try:
            with open(os.path.join(DATA_FOLDER, "original_route_names.json"), 'r', encoding='utf-8') as f:
                original_names_data = json.load(f)
                original_names = original_names_data.get("routes", [])
        except Exception as e:
            print(f"⚠️ Не удалось загрузить original_route_names.json: {e}")
            original_names = []

        # Сохраняем введенные пользовательские значения из текущего self.summary
        saved_user_input = {}
        for norm_name, data in self.summary.items():
            saved_user_input[norm_name] = {
                "report_distance": data.get("report_distance"),
                "report_duration_hours": data.get("report_duration_hours"),
                "report_duration_minutes": data.get("report_duration_minutes"),
                "comment": data.get("comment", "")
            }

        # Очищаем текущий summary перед заполнением актуальными данными
        current_summary = {} # Используем временный словарь для новых данных
        print("--- Обновление данных для сводки --- ")

        # Обрабатываем каждый маршрут из списка актуальных
        for original_name in original_names:
            sanitized_name = sanitize_filename(original_name)
            if not sanitized_name: continue
            
            print(f"  Обработка: {original_name}...")
            try:
                # Запрашиваем свежие данные маршрута (расстояние, точки и т.д.)
                route_data_response = get_route_data_endpoint(original_name)
                
                if route_data_response and not route_data_response.get("error"):
                    distance_data = route_data_response.get("distance_data", {})
                    distance_km = distance_data.get("total_distance") # Ожидаем км
                    duration_sec = distance_data.get("total_duration") # Ожидаем сек
                    duration_formatted = distance_data.get("formatted_duration", "Н/Д")
                    number_of_stops = route_data_response.get("number_of_stops", 0)
                    
                    # --- ПОЛУЧЕНИЕ ФИО ВОДИТЕЛЯ --- 
                    driver_name = self.drivers.get(sanitized_name, "—") # Получаем из self.drivers, дефолт "—"
                    print(f"    -> Водитель для {original_name}: {driver_name}")
                    # -----------------------------------
                    
                    # Создаем запись в НОВОЙ сводке
                    current_summary[sanitized_name] = {
                        "original_name": original_name,
                        "driver_name": driver_name, # <--- ДОБАВЛЕНО ПОЛЕ
                        "distance": distance_km if distance_km is not None else "Н/Д",
                        "duration_seconds": duration_sec,
                        "duration_formatted": duration_formatted,
                        "number_of_stops": number_of_stops, # Используем актуальное кол-во
                        "report_distance": None, # Будут восстановлены ниже
                        "report_duration_hours": None,
                        "report_duration_minutes": None,
                        "distance_difference": None,
                        "time_difference_formatted": None,
                        "time_difference_seconds": None, # <-- ДОБАВЛЕНО
                        "total_route_time_seconds": None, # Будут рассчитаны ниже
                        "total_route_time_formatted": None
                    }

                    # Восстанавливаем пользовательские данные (сохраненные выше)
                    if sanitized_name in saved_user_input:
                        for field in ["report_distance", "report_duration_hours", "report_duration_minutes", "comment"]:
                            if field in saved_user_input[sanitized_name]:
                                current_summary[sanitized_name][field] = saved_user_input[sanitized_name][field]
                        
                    # --- ПРИНУДИТЕЛЬНЫЙ ПЕРЕСЧЕТ ПОЛЕЙ ДЛЯ КАЖДОЙ ЗАПИСИ --- 
                    self._recalculate_summary_fields(sanitized_name, summary_dict=current_summary) # Передаем словарь для расчета
                        
                    print(f"    ✅ Данные для сводки получены и пересчитаны: {original_name}")
                else:
                    error_msg = route_data_response.get('message') if route_data_response else 'Данные не получены'
                    print(f"    ⚠️ Ошибка получения данных для маршрута {original_name}: {error_msg}")
            except Exception as e:
                print(f"    ❌ Исключение при обработке маршрута {original_name} для сводки: {e}")
                # import traceback
                # traceback.print_exc()

        # Обновляем основной summary новыми данными и сохраняем
        self.summary = current_summary
        self.save_to_disk() 
        print("--- Обновление сводки завершено --- ")
        return self.summary

    def save_to_disk(self):
         data = {
             "routes": self.routes,
             "summary": self.summary,
             "drivers": self.drivers, 
             "current_file": self.current_file,
             "global_service_time_minutes": self.global_service_time_minutes,
             "report_date_str": self.report_date_str # <--- СОХРАНЯЕМ ДАТУ
         }
         # Очищаем данные перед сохранением (убедимся, что функция работает корректно)
         # sanitized_data = sanitize_data_for_json(data) # Эта функция не нужна, json.dump сам справится с None
         with open(os.path.join(DATA_FOLDER, "route_data.json"), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2) # Используем data напрямую

    def load_from_disk(self):
        try:
            with open(os.path.join(DATA_FOLDER, "route_data.json"), "r", encoding="utf-8") as f:
                data = json.load(f)
                self.routes = data.get("routes", {})
                self.summary = data.get("summary", {})
                self.drivers = data.get("drivers", {}) 
                self.current_file = data.get("current_file", "")
                # Загружаем глобальное время, или 0 если нет
                self.global_service_time_minutes = data.get("global_service_time_minutes", 0) 
                self.report_date_str = data.get("report_date_str", None) # <--- ЗАГРУЖАЕМ ДАТУ

                print(f"💾 Данные загружены. Маршрутов: {len(self.routes)}, Записей в сводке: {len(self.summary)}, Водителей: {len(self.drivers)}, Время на точку: {self.global_service_time_minutes} мин, Дата отчета: {self.report_date_str}") # <-- Добавлен вывод даты

                # Миграция старых данных (добавляем новые поля в summary, если их нет)
                needs_recalculation = False
                for key, summary_item in self.summary.items():
                    if summary_item.setdefault("number_of_stops", None) is None:
                         # Пытаемся восстановить кол-во точек из routes, если возможно
                         route_entry = self.routes.get(key)
                         if route_entry and "route_points" in route_entry:
                             summary_item["number_of_stops"] = len(route_entry["route_points"])
                             needs_recalculation = True
                             print(f"   Восстановлено кол-во точек для {key}: {summary_item['number_of_stops']}")
                         else:
                             summary_item["number_of_stops"] = 0 # Или ставим 0, если не нашли
                             print(f"   Не удалось восстановить кол-во точек для {key}, установлено 0")
                    # Добавляем остальные поля со значением None, если их нет
                    if summary_item.setdefault("total_route_time_seconds", None) is None: needs_recalculation = True
                    if summary_item.setdefault("total_route_time_formatted", None) is None: needs_recalculation = True
                    # Убираем старое поле, если оно есть (на всякий случай)
                    summary_item.pop("service_time_per_stop_minutes", None)

                    # Убедимся, что обязательные поля для расчета существуют
                    summary_item.setdefault("report_duration_hours", None)
                    summary_item.setdefault("report_duration_minutes", None)
                    summary_item.setdefault("time_difference_formatted", None)
                    summary_item.setdefault("duration_seconds", None)
                    summary_item.setdefault("distance_difference", None)


                # Если добавили новые поля или восстановили кол-во точек, пересчитаем все
                if needs_recalculation:
                    print("   Некоторые поля отсутствовали или были восстановлены, запускаем полный пересчет сводки...")
                    for key in list(self.summary.keys()): # Используем list, чтобы не менять словарь во время итерации
                        self._recalculate_summary_fields(key)
                    self.save_to_disk() # Сохраняем пересчитанные данные

                return True
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"⚠️ Не удалось загрузить route_data.json: {e}. Начинаем с пустых данных.")
            self.routes = {}
            self.summary = {}
            self.current_file = ""
            self.global_service_time_minutes = 0 # Сброс глобального времени
            return False
        except Exception as e: # Ловим другие возможные ошибки при загрузке/миграции
             print(f"❌ Непредвиденная ошибка при загрузке/миграции данных route_data.json: {e}")
             # import traceback # Раскомментировать для детальной отладки
             # traceback.print_exc()
             self.routes = {}
             self.summary = {}
             self.current_file = ""
             self.global_service_time_minutes = 0
             return False

    # <--- НОВЫЙ МЕТОД --->
    def set_global_service_time(self, minutes):
        """Устанавливает глобальное время на точку и сохраняет его."""
        try:
            # Преобразуем в int, если не None и не пустая строка, иначе 0
            new_time = int(minutes) if minutes is not None and str(minutes).strip() != "" else 0
            if new_time < 0: new_time = 0 # Время не может быть отрицательным

            if new_time != self.global_service_time_minutes:
                print(f"🔄 Установка нового глобального времени на точку: {new_time} мин.")
                self.global_service_time_minutes = new_time
                # УБРАН НЕМЕДЛЕННЫЙ ПЕРЕСЧЕТ
                # recalculated_count = 0
                # for key in list(self.summary.keys()):
                #     self._recalculate_summary_fields(key)
                #     recalculated_count += 1
                # print(f"   Пересчитано маршрутов: {recalculated_count}")
                self.save_to_disk() # Сохраняем изменения
                return True
            else:
                print(f"ℹ️ Глобальное время на точку не изменилось ({new_time} мин).")
                return False
        except (ValueError, TypeError):
            print(f"⚠️ Некорректное значение для времени на точку: {minutes}. Оставляем текущее: {self.global_service_time_minutes} мин.")
            return False

    def get_office_coords(self):
        for route_name, data in self.routes.items():
            if data.get("geocoder_output"):
                for point in data["geocoder_output"]:
                    if point.get("type") == "office":
                        return {"lat": point["lat"], "lon": point["lon"]}
        return None

# Создаем глобальный объект для хранения данных
route_data = RouteData()
# Пробуем загрузить данные из файла
if not route_data.load_from_disk():
    print("💾 Инициализирован пустой объект RouteData.")

@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    # Принудительно обрабатываем маршрут "Липецк 2" при первом запуске
    # Закомментировано для предотвращения автоматической обработки
    # if not os.path.exists(os.path.join(os.path.dirname(BASE_DIR), "geocoded_results_Липецк_2.csv")):
    #     print("⚠️ Принудительно обрабатываем маршрут 'Липецк 2'...")
    #     process_route("Липецк 2")
    
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/api/health")
def health_check():
    return {"status": "ok"}

@app.get("/api/routes")
def get_routes():
    try:
        with open(os.path.join(DATA_FOLDER, "original_route_names.json"), 'r', encoding='utf-8') as f:
            original_names_data = json.load(f)
            return {"routes": sorted(original_names_data.get("routes", []))}
    except Exception as e:
        print(f"⚠️ Не удалось загрузить original_route_names.json: {e}")
        return {"routes": []}

# Функция для обработки отдельного маршрута
def process_route(original_route_name, report_date_str=None):
    """Обработка маршрута с сохранением оригинального имени"""
    # --- Добавлен лог полученной даты --- 
    print(f"-- [process_route] Запуск для '{original_route_name}'. Дата отчета: '{report_date_str}'")
    # --- КОНЕЦ ЛОГА ---
    process_success = False # Флаг общего успеха
    try:
        sanitized_name = sanitize_filename(original_route_name)
        parsed_addresses_file = os.path.join(PARSED_ADDRESSES_FOLDER, f"parsed_addresses_{sanitized_name}.csv")
        geocoded_file_path = os.path.join(GEOCODED_RESULTS_FOLDER, f"{sanitized_name}_geocoded.json")

        if not os.path.exists(parsed_addresses_file):
            print(f"❌ Файл {parsed_addresses_file} не найден. Невозможно геокодировать.")
            return False

        # Выполняем геокодирование
        geocoding_success = run_geocoding(parsed_addresses_file, geocoded_file_path, original_route_name)

        if not geocoding_success:
            print(f"❌ Геокодирование для {original_route_name} завершилось с ошибкой. Обработка маршрута прервана.")
            return False # Прерываем обработку, если геокодирование не удалось

        # Выполняем расчет расстояний ТОЛЬКО если геокодирование успешно
        # (Этот блок теперь точно будет выполняться только при успехе)
        calculation_success = False # Флаг успеха расчета расстояний
        if geocoding_success: # Запускаем расчет, только если геокодирование прошло
            print(f"\n▶️ Запуск расчета расстояний для {original_route_name}...")
            # --- ИЗМЕНЕНО: Передаем report_date_str в calculate_and_save_route ---
            calculation_success = route_distance.calculate_and_save_route(
                route_name=original_route_name, # Передаем оригинальное имя
                geocoded_file_path=geocoded_file_path, 
                traffic_mode='statistics', # <-- ЯВНО УКАЗЫВАЕМ РЕЖИМ
                report_date_str=report_date_str 
            )
            # --- КОНЕЦ ИЗМЕНЕНИЯ --- 
        else:
             print(f"ℹ️ Пропуск расчета расстояний для {original_route_name}, т.к. геокодирование не удалось.")
        
        if calculation_success:
            print(f"✅ Расчет расстояний для {original_route_name} завершен успешно (через функцию).")
        else:
            print(f"❌ Ошибка при расчете расстояний для {original_route_name} (через функцию).")
            # Можно решить, прерывать ли весь процесс или нет
            # return False # Прервать, если расчет расстояний критичен

        # Добавление в route_data
        if calculation_success: # Добавляем только если расчет прошел
            try:
                # Получаем данные для добавления в RouteData (включая distance_data)
                # route_info теперь будет содержать и geocoded_data и distance_data
                # Вызываем функцию, которая ТОЛЬКО читает файлы и форматирует
                route_info = get_route_data_endpoint(original_route_name)
                if route_info and not route_info.get("error"):
                    route_data.add_route(original_route_name, route_info)
                    route_data.save_to_disk()
                    print(f"✅ Маршрут {original_route_name} добавлен в сводку")
                    process_success = True # Весь процесс успешен
                else:
                    error_msg = route_info.get("message", "Не удалось получить данные маршрута") if route_info else "Не удалось получить данные маршрута"
                    print(f"⚠️ Не удалось добавить маршрут {original_route_name} в сводку: {error_msg}")
            except Exception as e:
                print(f"❌ Ошибка при добавлении маршрута {original_route_name} в сводку: {str(e)}")
        else:
             print(f"ℹ️ Маршрут {original_route_name} не добавлен в сводку, т.к. расчет расстояний не удался.")

    except Exception as e_proc:
        print(f"❌ Непредвиденная ошибка в process_route для '{original_route_name}': {e_proc}")
        import traceback
        traceback.print_exc()
        process_success = False
        
    return process_success

# --- КОНЕЦ ФУНКЦИИ process_route ---

@app.get("/api/route-data/{route_name}")
def get_route_data_endpoint(route_name: str = Path(...)):
    original_route_name = route_name
    if not original_route_name:
        raise HTTPException(status_code=400, detail="Имя маршрута не указано")

    try:
        logging.info(f"Getting route data for '{original_route_name}'")
        file_name = sanitize_filename(original_route_name)
        # --- ИЗМЕНЕНО: Определяем пути к обоим форматам ---
        geocoded_json_file = os.path.join(GEOCODED_RESULTS_FOLDER, f"{file_name}_geocoded.json")
        geocoded_csv_file = os.path.join(GEOCODED_RESULTS_FOLDER, f"geocoded_results_{file_name}.csv")
        distance_data_json_file = os.path.join(ROUTE_RESULTS_FOLDER, f"{file_name}_distance_data.json") # Приоритет для пересчитанных
        route_results_json_file = os.path.join(ROUTE_RESULTS_FOLDER, f"route_results_{file_name}.json") # Старый файл от process_route
        # --- КОНЕЦ ИЗМЕНЕНИЯ ---
        summary_item = route_data.summary.get(file_name, {}) # Получаем заранее

        # --- ИЗМЕНЕНО: Чтение файла с результатами расстояний (приоритет _distance_data.json) ---
        distance_data = {}
        distance_file_to_read = None

        # --- [DEBUG get_route_data] Лог A: Проверка существования файлов перед чтением --- 
        print(f"--- [DEBUG get_route_data Log A] Route: {original_route_name}")
        print(f"    Checking exists ({os.path.basename(distance_data_json_file)}): {os.path.exists(distance_data_json_file)}")
        print(f"    Checking exists ({os.path.basename(route_results_json_file)}): {os.path.exists(route_results_json_file)}")
        # --- КОНЕЦ ЛОГА A ---

        if os.path.exists(distance_data_json_file):
            distance_file_to_read = distance_data_json_file
            logging.info(f"Found distance data file: {distance_file_to_read}")
        elif os.path.exists(route_results_json_file):
            distance_file_to_read = route_results_json_file
            logging.warning(f"Distance data file not found, falling back to route results: {distance_file_to_read}")
        else:
            logging.warning(f"Neither distance_data nor route_results file found for route {original_route_name}")
            distance_data = {"error": True, "error_message": f"Файлы с данными о расстоянии не найдены для {original_route_name}"}

        if distance_file_to_read:
            try:
                with open(distance_file_to_read, 'r', encoding='utf-8') as f:
                    distance_data_loaded = json.load(f) # Загружаем данные

                # --- ВАЖНО: Копируем загруженные данные в distance_data ---
                # Это необходимо, чтобы дальнейшая логика форматирования работала
                # с правильным объектом, даже если он пустой.
                if isinstance(distance_data_loaded, dict):
                     distance_data = distance_data_loaded.copy()
                else:
                     # Если формат не словарь, логируем ошибку и используем пустой словарь
                     logging.error(f"Invalid format in distance file {distance_file_to_read}. Expected dict, got {type(distance_data_loaded)}.")
                     distance_data = {}

                # Форматируем данные для фронтенда (остальная логика без изменений)
                if 'total_distance' in distance_data and distance_data['total_distance'] is not None:
                    try:
                        # --- ИЗМЕНЕНО: Обработка как КМ, так и метров ---
                        # total_distance может быть уже в КМ (округленных) или в метрах
                        distance_val = float(distance_data['total_distance'])
                        # Если значение большое (>10000), скорее всего это метры, конвертируем и округляем
                        if distance_val > 10000:
                             distance_km_rounded = round(distance_val / 1000)
                             print(f"   [Dist Format] Detected meters ({distance_val}), converting to rounded km: {distance_km_rounded}")
                        else:
                             # Иначе считаем, что это уже округленные КМ
                             distance_km_rounded = int(distance_val)
                             print(f"   [Dist Format] Detected km ({distance_val}), using as int: {distance_km_rounded}")

                        distance_data['formatted_distance'] = f"{distance_km_rounded} км"
                        # Сохраняем округленные КМ для консистентности
                        distance_data['total_distance'] = distance_km_rounded
                    except (ValueError, TypeError) as e_dist:
                        print(f"⚠️ Не удалось преобразовать total_distance '{distance_data['total_distance']}': {e_dist}")
                        distance_data['formatted_distance'] = "Ошибка"
                        distance_data['total_distance'] = None
                else:
                    distance_data['formatted_distance'] = "Н/Д"
                    distance_data['total_distance'] = None # Убедимся, что None, если ключа нет

                if 'total_duration' in distance_data and distance_data['total_duration'] is not None:
                    try:
                        total_seconds = int(distance_data['total_duration'])
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        formatted_duration = f"{hours} ч {minutes} мин" if hours > 0 else f"{minutes} мин"
                        # Коррекция для 0 секунд
                        if total_seconds == 0: formatted_duration = "0 мин"
                        distance_data['formatted_duration'] = formatted_duration
                        # Убедимся, что total_duration остается в секундах
                        distance_data['total_duration'] = total_seconds
                    except (ValueError, TypeError) as e_dur:
                         print(f"⚠️ Не удалось преобразовать total_duration '{distance_data['total_duration']}': {e_dur}")
                         distance_data['formatted_duration'] = "Ошибка"
                         distance_data['total_duration'] = None
                else:
                    distance_data['formatted_duration'] = "Н/Д"
                    distance_data['total_duration'] = None # Убедимся, что None

            except Exception as e:
                logging.error(f"Error reading distance/route results file {distance_file_to_read} for {original_route_name}: {e}")
                distance_data = {"error": True, "error_message": f"Ошибка чтения файла результатов: {str(e)}"}
        # --- КОНЕЦ ИЗМЕНЕНИЯ ЧТЕНИЯ ФАЙЛА РАССТОЯНИЙ ---

        # --- ИЗМЕНЕНО: Чтение geocoded_data (приоритет JSON) ---
        geocoded_data = []
        route_points = [] # Будет содержать только точки с валидными lat/lon
        # number_of_stops_actual = 0 # Старая переменная больше не нужна в этом месте
        geocoded_data_source = None # 'json' или 'csv'
        # visible_points_count = 0 # Инициализация удалена, будет рассчитываться позже
        points_for_summary = 0   # Инициализация

        try:
            # 1. Пытаемся прочитать JSON
            if os.path.exists(geocoded_json_file):
                logging.info(f"Attempting to load geocoded data from JSON: {geocoded_json_file}")
                try:
                    with open(geocoded_json_file, 'r', encoding='utf-8') as f_json:
                        loaded_data = json.load(f_json)
                        if isinstance(loaded_data, list) and all(isinstance(item, dict) for item in loaded_data):
                            geocoded_data = loaded_data # Используем данные из JSON
                            geocoded_data_source = 'json'
                            print(f"   Successfully loaded {len(geocoded_data)} points from JSON.")
                        else:
                            logging.error(f"Invalid format in {geocoded_json_file}. Expected list of dicts.")
                except Exception as e_json:
                    logging.error(f"Error reading {geocoded_json_file}: {e_json}. Falling back to CSV.")

            # 2. Если JSON не прочитан, пытаемся прочитать CSV
            if not geocoded_data and os.path.exists(geocoded_csv_file):
                logging.warning(f"JSON not found or invalid. Attempting to load from CSV: {geocoded_csv_file}")
                try:
                    geocoded_df = pd.read_csv(geocoded_csv_file)
                    geocoded_df.replace([np.nan, np.inf, -np.inf], None, inplace=True) # Заменяем NaN/inf на None
                    geocoded_data = geocoded_df.to_dict('records') # Конвертируем в список словарей
                    geocoded_data_source = 'csv'
                    print(f"   Successfully loaded {len(geocoded_data)} points from CSV.")
                    # --- Опционально: Конвертировать и сохранить CSV в JSON ---
                    # try:
                    #     with open(geocoded_json_file, 'w', encoding='utf-8') as f_conv:
                    #         json.dump(geocoded_data, f_conv, ensure_ascii=False, indent=4)
                    #     print(f"   Successfully converted and saved data to {geocoded_json_file}")
                    # except Exception as e_conv:
                    #     logging.warning(f"Could not save converted CSV data to JSON ({geocoded_json_file}): {e_conv}")
                    # --- Конец опциональной конвертации ---
                except Exception as e_csv:
                    logging.error(f"Error reading or processing CSV file {geocoded_csv_file}: {e_csv}")
                    geocoded_data = [] # Оставляем пустым при ошибке CSV

            # 3. Если нет ни JSON, ни CSV
            elif not geocoded_data:
                logging.warning(f"Could not find geocoded file (JSON or CSV) for route {original_route_name}")
                
            # --- ЛОГ 1: Содержимое geocoded_data после чтения (УДАЛЕН) --- 
            # print(f"[DEBUG get_route_data] After read: geocoded_data length={len(geocoded_data)}")
            # if len(geocoded_data) >= 1:
            #     print(f"  First point: {geocoded_data[0].get('excel_row')}, Input: {geocoded_data[0].get('input')}")
            # if len(geocoded_data) >= 2:
            #     print(f"  Last point: {geocoded_data[-1].get('excel_row')}, Input: {geocoded_data[-1].get('input')}")
            # --- КОНЕЦ УДАЛЕНИЯ ЛОГА 1 ---

            # 4. Обработка прочитанных geocoded_data (если они есть)
            if geocoded_data:
                 for point_data in geocoded_data:
                     # Добавляем в route_points только если есть валидные координаты
                     if 'lon' in point_data and 'lat' in point_data and point_data['lon'] is not None and point_data['lat'] is not None:
                        try:
                             lat = float(point_data['lat'])
                             lon = float(point_data['lon'])
                             if -90 <= lat <= 90 and -180 <= lon <= 180:
                                route_points.append({
                                     'lon': lon,
                                     'lat': lat,
                                     # Используем description для адреса на карте, input для исходного
                                     'address': point_data.get('description', point_data.get('input', '')),
                                     'original_address': point_data.get('input', '')
                                 })
                             else: # Логгируем некорректный диапазон
                                 print(f"⚠️ Invalid coordinate range in geocoded_data for {original_route_name}: lat={lat}, lon={lon}")
                        except (ValueError, TypeError):
                             print(f"⚠️ Invalid coordinate types in geocoded_data for {original_route_name}: {point_data.get('lat')}, {point_data.get('lon')}")

                 # --- УДАЛЕН БЛОК НЕПРАВИЛЬНОГО ПОДСЧЕТА И ОБНОВЛЕНИЯ SUMMARY ---
                 # # Считаем точки с координатами (без офиса, т.к. он еще не добавлен)
                 # number_of_stops_actual = len(route_points)
                 #
                 # # Обновляем количество точек в summary, если оно расходится 
                 # if file_name in route_data.summary:
                 #     current_summary_stops = route_data.summary[file_name].get("number_of_stops")
                 #     # --- ЛОГ 6: Перед сравнением --- 
                 #     print(f"[DEBUG get_route_data] Comparing points: current_in_summary={current_summary_stops}, calculated_now={number_of_stops_actual}")
                 #     # --- КОНЕЦ ЛОГА 6 ---
                 #     if current_summary_stops is None or current_summary_stops != number_of_stops_actual:
                 #         print(f"⚠️ Обновление кол-ва точек для {original_route_name} с {current_summary_stops} на {number_of_stops_actual}")
                 #         route_data.summary[file_name]["number_of_stops"] = number_of_stops_actual
                 #         route_data._recalculate_summary_fields(file_name)
                 #         route_data.save_to_disk() # Сохраняем изменение
                 #         summary_item = route_data.summary.get(file_name, {}) # Обновляем локальный summary_item
                 #     # --- КОНЕЦ ИЗМЕНЕНИЯ ---
                 # else:
                 #     logging.warning(f"Route {original_route_name} (sanitized: {file_name}) not found in summary data.")
                 # --- КОНЕЦ УДАЛЕННОГО БЛОКА ---
            else: # Если geocoded_data пуст
                 # number_of_stops_actual = 0 # Больше не нужно здесь
                 route_points = []

        except Exception as e:
             # Ловим общую ошибку чтения/обработки геокодированных данных
             logging.error(f"Error processing geocoding data for {original_route_name}: {e}")
             geocoded_data = []
             route_points = []
             number_of_stops_actual = 0
        # --- КОНЕЦ ИЗМЕНЕНИЯ ЧТЕНИЯ GEOCODED_DATA ---

        # --- ДОБАВЛЕНИЕ ОФИСА РТК В НАЧАЛО И КОНЕЦ СПИСКОВ ДЛЯ ФРОНТЕНДА ---
        # (Эта логика остается без изменений, т.к. она работает с уже прочитанными geocoded_data и route_points)
        try:
            office_loc = config.OFFICE_LOCATION
            if office_loc and "lat" in office_loc and "lon" in office_loc:
                office_lat = office_loc["lat"]
                office_lon = office_loc["lon"]
                office_name = office_loc.get("name", "Офис РТК")

                # Представление для таблицы geocoded_data
                office_geocoded_repr = {
                    "excel_row": "СТАРТ",
                    "route_name": original_route_name,
                    "input": office_name,
                    "found": True,
                    "type": "office",
                    "description": office_name,
                    "lat": office_lat,
                    "lon": office_lon,
                    "error": None
                }
                # Представление для карты route_points
                office_route_point_repr = {
                    "lat": office_lat,
                    "lon": office_lon,
                    "address": office_name,
                    "original_address": office_name
                }

                # Добавляем в начало
                geocoded_data.insert(0, office_geocoded_repr)
                route_points.insert(0, office_route_point_repr)

                # Добавляем в конец (копируем, чтобы excel_row можно было изменить)
                office_geocoded_repr_end = office_geocoded_repr.copy()
                office_geocoded_repr_end["excel_row"] = "ФИНИШ"
                geocoded_data.append(office_geocoded_repr_end)
                route_points.append(office_route_point_repr) # Координаты те же

                print(f"✅ Добавлен офис РТК в начало/конец данных для маршрута {original_route_name}")

            else:
                print("⚠️ Координаты OFFICE_LOCATION не найдены или неполны в config.py. Офис не добавлен.")
        except Exception as e_office:
            print(f"❌ Ошибка при добавлении офиса РТК: {e_office}")
        # --- КОНЕЦ ДОБАВЛЕНИЯ ОФИСА ---

        # --- Добавление офиса (точек СТАРТ/ФИНИШ) --- 
        office_coords = route_data.get_office_coords()
        if office_coords:
            start_point = {
                "excel_row": "СТАРТ",
                "input": "Офис (Старт)",
                "found": "Офис (Старт)",
                "lat": office_coords['lat'],
                "lon": office_coords['lon'],
                "type": "office",
                "description": "Офис (Старт)",
                "hidden": False # Офис по умолчанию видимый
            }
            end_point = {
                "excel_row": "ФИНИШ",
                "input": "Офис (Финиш)",
                "found": "Офис (Финиш)",
                "lat": office_coords['lat'],
                "lon": office_coords['lon'],
                "type": "office",
                "description": "Офис (Финиш)",
                "hidden": False # Офис по умолчанию видимый
            }
            # Вставляем старт в начало, только если его там еще нет
            # --- ЛОГ 2: Перед проверкой СТАРТ (УДАЛЕН) --- 
            start_check_condition = not geocoded_data or geocoded_data[0].get('excel_row') != 'СТАРТ'
            # print(f"[DEBUG get_route_data] Before START check: condition is {start_check_condition}")
            # --- КОНЕЦ УДАЛЕНИЯ ЛОГА 2 ---
            if start_check_condition:
                geocoded_data.insert(0, start_point)
            # Добавляем финиш в конец, только если его там еще нет
            # --- ЛОГ 3: Перед проверкой ФИНИШ (УДАЛЕН) --- 
            finish_check_condition = not geocoded_data or geocoded_data[-1].get('excel_row') != 'ФИНИШ'
            # print(f"[DEBUG get_route_data] Before FINISH check: condition is {finish_check_condition}")
            # --- КОНЕЦ УДАЛЕНИЯ ЛОГА 3 ---
            if finish_check_condition:
                geocoded_data.append(end_point)
            # --- ЛОГ 4: После добавления офиса (УДАЛЕН) --- 
            # print(f"[DEBUG get_route_data] After office check: final geocoded_data length={len(geocoded_data)}")
            # if len(geocoded_data) >= 1: print(f"  Final first point: {geocoded_data[0].get('excel_row')}, Input: {geocoded_data[0].get('input')}")
            # if len(geocoded_data) >= 2: print(f"  Final last point: {geocoded_data[-1].get('excel_row')}, Input: {geocoded_data[-1].get('input')}")
            # --- КОНЕЦ УДАЛЕНИЯ ЛОГА 4 ---
        # --- Конец Добавления офиса ---
        
        # --- НОВЫЙ БЛОК: Пересчитываем количество точек и обновляем summary ПОСЛЕ добавления/проверки офиса ---
        # Подсчет видимых промежуточных точек (не СТАРТ, не ФИНИШ, не hidden)
        points_for_summary = sum(
            1 for p in geocoded_data 
            if not p.get('hidden', False) 
            and p.get('excel_row') not in ['СТАРТ', 'ФИНИШ']
        )
        
        # Обновляем количество точек в summary, если оно расходится 
        if file_name in route_data.summary:
            current_summary_stops = route_data.summary[file_name].get("number_of_stops")
            # --- ЛОГ 6: Перед сравнением (УДАЛЕН) --- 
            # print(f"[DEBUG get_route_data] Comparing points after office: current_in_summary={current_summary_stops}, calculated_now={points_for_summary}")
            # --- КОНЕЦ УДАЛЕНИЯ ЛОГА 6 ---
            if current_summary_stops is None or current_summary_stops != points_for_summary:
                # --- ЛОГ 7: Если происходит обновление (УДАЛЕН) --- 
                print(f"ℹ️ Обновление кол-ва промежуточных точек для '{original_route_name}' с {current_summary_stops} на {points_for_summary} в summary.")
                # --- КОНЕЦ УДАЛЕНИЯ ЛОГА 7 ---
                route_data.summary[file_name]["number_of_stops"] = points_for_summary
                # Пересчитываем зависимые поля в summary (время на маршруте и т.д.)
                route_data._recalculate_summary_fields(file_name) 
                # Сохраняем изменения в route_data.json
                route_data.save_to_disk() 
                # Обновляем локальный summary_item для возврата актуальных данных
                summary_item = route_data.summary.get(file_name, {}) 
        else:
            # Логируем, если маршрут не найден в summary (хотя это маловероятно на данном этапе)
            logging.warning(f"Route {original_route_name} (sanitized: {file_name}) not found in summary data during post-office update.")
        # --- КОНЕЦ НОВОГО БЛОКА ---

        # --- Старый блок подсчета (УДАЛЕН) ---
        # # --- Пересчитываем количество точек ПОСЛЕ добавления офиса ---
        # visible_points_count = sum(1 for p in geocoded_data if not p.get('hidden', False))
        # points_for_summary = max(0, visible_points_count - 2)
        # # --- ЛОГ 5: Результаты расчета --- 
        # print(f"[DEBUG get_route_data] Point calculation: visible_count={visible_points_count}, points_for_summary={points_for_summary}")
        # # --- КОНЕЦ ЛОГА 5 ---
        # 
        # # Обновляем количество точек в summary, если оно расходится 
        # if file_name in route_data.summary:
        #     current_summary_stops = route_data.summary[file_name].get("number_of_stops")
        #     # --- ЛОГ 6: Перед сравнением --- 
        #     print(f"[DEBUG get_route_data] Comparing points: current_in_summary={current_summary_stops}, calculated_now={points_for_summary}")
        #     # --- КОНЕЦ ЛОГА 6 ---
        #     if current_summary_stops is None or current_summary_stops != points_for_summary:
        #         # --- ЛОГ 7: Если происходит обновление --- 
        #         print(f"[DEBUG get_route_data] Updating summary stops from {current_summary_stops} to {points_for_summary}")
        #         # --- КОНЕЦ ЛОГА 7 ---
        #         route_data.summary[file_name]["number_of_stops"] = points_for_summary
        #         # ... (остальной код) ...
        # --- КОНЕЦ УДАЛЕНИЯ СТАРОГО БЛОКА ---

        # Возвращаем данные, включая оригинальное имя маршрута и новые поля
        # Используем number_of_stops из актуального (возможно, только что обновленного) summary_item
        return {
            "route_name": original_route_name,
            "geocoder_output": geocoded_data, # <--- Теперь с офисом и правильной структурой
            "route_points": route_points,     # <--- Теперь с офисом
            "distance_data": distance_data,
            "number_of_stops": summary_item.get("number_of_stops", 0), # Берем актуальное значение из summary
            "total_route_time_formatted": summary_item.get("total_route_time_formatted", "Н/Д"),
            "global_service_time_minutes": route_data.global_service_time_minutes
        }
    except Exception as e:
        logging.error(f"Error in get_route_data_endpoint for {original_route_name}: {e}")
        # import traceback
        # traceback.print_exc()
        return {"error": True, "message": str(e), "route_name": original_route_name}

@app.get("/api/summary")
def get_summary_endpoint():
    summary = route_data.get_summary() # Этот метод теперь сам пересчитывает и сохраняет
    result = []
    # Перебираем актуальный route_data.summary после вызова get_summary()
    for norm_name, data in route_data.summary.items():
        result.append({
            "route_name": data.get("original_name", norm_name),
            "driver_name": data.get("driver_name", "—"),
            "distance": data.get("distance", "Н/Д"),
            "duration_seconds": data.get("duration_seconds"), # Время в пути (сек)
            "duration": data.get("duration_formatted", "Н/Д"), # Время в пути (формат)
            "report_distance": data.get("report_distance"),
            "report_duration_hours": data.get("report_duration_hours"),
            "report_duration_minutes": data.get("report_duration_minutes"),
            "distance_difference": data.get("distance_difference"),
            "time_difference_formatted": data.get("time_difference_formatted"),
            "time_difference_seconds": data.get("time_difference_seconds"), # <-- ДОБАВЛЕНО
            "total_route_time_seconds": data.get("total_route_time_seconds"), # <-- Добавляем общее время в секундах
            "total_route_time_formatted": data.get("total_route_time_formatted", "Н/Д"), # Время на маршруте (формат)
            "number_of_stops": data.get("number_of_stops", "Н/Д"), # <--- ДОБАВЛЕНО КОЛ-ВО ТОЧЕК
            "comment": data.get("comment", "") # Комментарий
        })
    # Добавляем глобальное время на точку в ответ
    return {
        "summary": result, # <<< УБРАНА СОРТИРОВКА
        "global_service_time_minutes": route_data.global_service_time_minutes
     }

@app.post("/api/upload")
async def upload_excel(file: UploadFile, 
                     service_time_per_stop_minutes: int = Form(0),
                     report_date: str = Form("")): # <--- Параметр даты уже есть
    save_path = os.path.join(UPLOAD_FOLDER, file.filename)
    with open(save_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # Сохраняем имя текущего файла
    route_data.current_file = file.filename
    # Сохраняем дату отчета
    route_data.report_date_str = report_date # <--- СОХРАНЕНИЕ ДАТЫ
    print(f"--- Received report date in /api/upload: {report_date} ---") # <-- Лог даты

    # --- Печатаем полученное значение времени ---
    print(f"--- Received time per stop in /api/upload: {service_time_per_stop_minutes} (type: {type(service_time_per_stop_minutes)}) ---")
    # --- Конец блока ---

    # --- Устанавливаем глобальное время на точку ---
    route_data.set_global_service_time(service_time_per_stop_minutes)
    # --- Конец блока ---

    try:
        # --- Инициализация переменных для временных файлов (ШАГ 3.1) --- 
        temp_upload_path = None # Исходный временный файл загрузки
        corrected_path = None   # Потенциально исправленный файл
        path_to_read = None     # Финальный путь для чтения
        # --- Конец инициализации ---
        
        # --- ШАГ 3.2 (REVISED): Сохраняем загруженный файл во временный файл --- 
        try:
            # Создаем имя временного файла
            temp_fd, temp_upload_path = tempfile.mkstemp(suffix=os.path.splitext(file.filename)[1])
            os.close(temp_fd) # Закрываем дескриптор, нам нужен только путь
            print(f"INFO: Создан временный путь: {temp_upload_path}")

            # Открываем файл для записи и копируем содержимое асинхронно
            bytes_written = 0
            with open(temp_upload_path, 'wb') as temp_f_out:
                # !!! Добавляем await file.seek(0) перед циклом чтения !!!
                await file.seek(0) 
                while True:
                    # Читаем асинхронно порцию данных из загружаемого файла
                    chunk = await file.read(8192) # 8KB chunk size
                    if not chunk:
                        break
                    temp_f_out.write(chunk)
                    bytes_written += len(chunk)
            # Файл temp_f_out закрывается здесь при выходе из with
            print(f"INFO: Файл сохранен во временный путь: {temp_upload_path} ({bytes_written} байт)")
            
            # Добавим проверку размера файла на всякий случай
            if bytes_written == 0:
                 raise ValueError("Сохраненный временный файл оказался пустым.")
                 
        except Exception as e_save:
             print(f"ERROR: Не удалось сохранить загруженный файл: {e_save}")
             # Попытка удалить временный файл, если он был создан частично
             if temp_upload_path and os.path.exists(temp_upload_path):
                 try: os.remove(temp_upload_path)
                 except Exception: pass
             raise HTTPException(status_code=500, detail=f"Ошибка сохранения файла: {e_save}")
        # --- Конец Шага 3.2 (REVISED) ---
        
        # --- ШАГ 3.3: Проверяем и исправляем временный файл --- 
        path_to_read = fix_xlsx_casing(temp_upload_path)
        if path_to_read is None:
            # fix_xlsx_casing вернула None, ошибка (напр. BadZipFile или другая при исправлении)
            if temp_upload_path and os.path.exists(temp_upload_path):
                 try: os.remove(temp_upload_path)
                 except Exception: pass
            # Теперь возвращаем ошибку 400, т.к. проблема скорее всего в файле
            raise HTTPException(status_code=400, detail="Ошибка проверки файла. Убедитесь, что это корректный XLSX файл.")
        elif path_to_read != temp_upload_path:
            corrected_path = path_to_read # Запоминаем путь к исправленному файлу
            print(f"INFO: Файл был исправлен, используется новый путь: {corrected_path}")
        else:
            print("INFO: Исправление файла не потребовалось.")
        # --- Конец Шага 3.3 --- 
        
        # --- Чтение Excel и определение маршрутов (ШАГ 3.4) --- 
        print(f"--- Чтение файла для обработки: {path_to_read} ---")
        df = pd.read_excel(path_to_read, engine='openpyxl')
        # --- Конец Шага 3.4 ---
        
        # --- Проверяем, что словарь листов не пустой (вместо df.empty) --- 
        if df.empty:
             raise ValueError("Excel файл не содержит листов или пуст после чтения.")
        # --- Конец проверки --- 

        all_routes_addresses = {} # Словарь: route_name -> [(row_num, address), ...]
        route_names = []          # Список имен маршрутов в порядке их появления
        current_route = None
        seen_kontragents_in_route = set() # <-- Множество для отслеживания контрагентов в ТЕКУЩЕМ маршруте
        
        # --- Добавлена проверка на пустой файл --- 
        if df.empty:
             raise ValueError("Загруженный Excel файл пуст.")
        # --- Конец проверки --- 

        # --- Ищем колонки динамически (аналогично parsing_route.py) --- 
        region_col_name = None
        address_col_name = "Адрес доставки"
        kontragent_col_name = None # <-- Имя колонки контрагентов
        driver_col_name = None # <--- Имя колонки водителя
        
        for col in df.columns:
            col_str = str(col).strip()
            if region_col_name is None and col_str.startswith("Регион/Маршрут"):
                region_col_name = col_str
            if kontragent_col_name is None and col_str.startswith("Контрагентов"):
                 kontragent_col_name = col_str
            if driver_col_name is None and col_str.startswith("Водитель"): # <--- Ищем колонку Водитель
                 driver_col_name = col_str
                
        # Проверяем наличие всех необходимых колонок
        missing_cols = []
        if region_col_name is None: missing_cols.append("колонка, начинающаяся с 'Регион/Маршрут'")
        if address_col_name not in df.columns: missing_cols.append(f"колонка '{address_col_name}'")
        if kontragent_col_name is None: missing_cols.append("колонка, начинающаяся с 'Контрагентов'")
        if driver_col_name is None: missing_cols.append("колонка, начинающаяся с 'Водитель'") # <--- Проверяем колонку Водитель
        
        if missing_cols:
             raise ValueError(f"Отсутствуют необходимые колонки: { ' и '.join(missing_cols) }.")
        print(f"  ✅ Найдены колонки: '{region_col_name}', '{address_col_name}', '{kontragent_col_name}', '{driver_col_name}'") # <--- Добавили в лог
        # --- Конец динамического поиска --- 
        
        # Сбрасываем старые данные о водителях перед обработкой нового файла
        route_data.drivers = {}
        print("  ℹ️ Словарь водителей очищен перед обработкой нового файла.")

        for idx, row in df.iterrows():
            # Используем найденные имена колонок
            region = row.get(region_col_name)
            address = row.get(address_col_name)
            kontragent = row.get(kontragent_col_name) # <-- Получаем контрагента
            driver = row.get(driver_col_name) # <--- Получаем водителя
            
            # Определяем начало нового маршрута
            is_new_route_marker = False
            if pd.notna(region) and isinstance(region, str) and region.strip():
                potential_new_route = region.strip()
                if pd.isna(address) or not str(address).strip() or current_route is None or potential_new_route != current_route:
                     is_new_route_marker = True
                     current_route = potential_new_route
                     if current_route not in route_names: # Сохраняем порядок только для НЕПУСТЫХ маршрутов
                         route_names.append(current_route)
                     if current_route not in all_routes_addresses:
                         all_routes_addresses[current_route] = []
                     seen_kontragents_in_route.clear() # <-- Очищаем сет контрагентов для нового маршрута
                     # driver_found_for_current_route = False # <-- Сбрасываем флаг для нового маршрута (Более простой способ - проверка наличия ключа)
            
            # Добавляем адрес к ТЕКУЩЕМУ маршруту, если:
            # 1. Маршрут определен
            # 2. Адрес есть
            # 3. Контрагент есть и он еще НЕ встречался в ЭТОМ маршруте
            if current_route and pd.notna(address) and isinstance(address, str) and address.strip():
                excel_row = idx + 2 # +1 за 0-индекс, +1 за строку заголовка
                if pd.notna(kontragent) and isinstance(kontragent, str) and kontragent.strip():
                    kontragent_key = kontragent.strip()
                    if kontragent_key not in seen_kontragents_in_route:
                        seen_kontragents_in_route.add(kontragent_key) # Добавляем контрагента в сет
                        all_routes_addresses[current_route].append((excel_row, address.strip()))
                        
                        # --- СОХРАНЕНИЕ ВОДИТЕЛЯ --- 
                        sanitized_route_name = sanitize_filename(current_route)
                        # Проверяем, что водитель еще не найден для этого маршрута И что значение водителя валидно
                        if sanitized_route_name and sanitized_route_name not in route_data.drivers:
                            if pd.notna(driver) and isinstance(driver, str) and driver.strip():
                                driver_name = driver.strip()
                                route_data.drivers[sanitized_route_name] = driver_name
                                print(f"    🧑 Найден водитель для '{current_route}': {driver_name}")
                        # --- КОНЕЦ СОХРАНЕНИЯ ВОДИТЕЛЯ ---
                        
                    # else: # Логгирование дубликата (опционально)
                        # print(f"  [upload] Пропуск строки {excel_row}: дубль контрагента '{kontragent_key}' в '{current_route}'")
                # else: # Логгирование отсутствующего контрагента (опционально)
                     # print(f"  [upload] Пропуск строки {excel_row}: нет контрагента в '{current_route}'")
        
        # Удаляем маршруты без адресов из обоих списков
        route_names = [r for r in route_names if all_routes_addresses.get(r)]
        all_routes_addresses = {r: adds for r, adds in all_routes_addresses.items() if adds}

        if not route_names:
            raise ValueError("Не найдено ни одного маршрута с адресами в файле.")
            
        print(f"✅ В файле найдено маршрутов с адресами: {len(route_names)}")
        
        # --- Сохраняем route_data с обновленными водителями ПЕРЕД запуском парсинга --- 
        if route_data.drivers: # Сохраняем, только если нашли хотя бы одного водителя
            print("💾 Сохраняем route_data с информацией о водителях...")
            route_data.save_to_disk()
        # -----------------------------------------------------------------------------
        
        # --- Очистка старых файлов и кеша для маршрутов из нового файла --- 
        print("\n🧹 Начинаем очистку старых данных для маршрутов из файла...")
        routes_cleaned_count = 0
        # Добавляем очистку summary CSV, если он существует
        summary_csv_path = os.path.join(SUMMARY_CSV_FOLDER, "current_summary.csv")
        # Очищаем его только если загрузка прошла успешно до этого момента
        # (Лучше очищать после успешной обработки всех маршрутов)
        # if os.path.exists(summary_csv_path):
        #     try:
        #         os.remove(summary_csv_path)
        #         print(f"  🗑️ Удален старый файл сводки: {os.path.basename(summary_csv_path)}")
        #     except OSError as e:
        #         print(f"  ❌ Ошибка при удалении файла {summary_csv_path}: {e}")

        for route_name_to_clean in route_names:
            file_name = sanitize_filename(route_name_to_clean)
            if not file_name:
                 continue
                 
            # Пути к файлам для удаления
            parsed_path = os.path.join(PARSED_ADDRESSES_FOLDER, f"parsed_addresses_{file_name}.csv")
            geocoded_path = os.path.join(GEOCODED_RESULTS_FOLDER, f"geocoded_results_{file_name}.csv")
            route_results_path = os.path.join(ROUTE_RESULTS_FOLDER, f"route_results_{file_name}.json")
            # --- ДОБАВЛЕНО: Пути к JSON файлам от пересчета ---
            geocoded_recalc_json_path = os.path.join(GEOCODED_RESULTS_FOLDER, f"{file_name}_geocoded.json")
            distance_recalc_json_path = os.path.join(ROUTE_RESULTS_FOLDER, f"{file_name}_distance_data.json")
            # --- КОНЕЦ ДОБАВЛЕНИЯ ---
            
            files_to_delete = [
                 parsed_path,
                 geocoded_path,
                 route_results_path,
                 geocoded_recalc_json_path, # Добавляем JSON для удаления
                 distance_recalc_json_path # Добавляем JSON для удаления
            ]
            deleted_something_for_route = False

            # Удаляем файлы
            for file_path in files_to_delete:
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        print(f"  🗑️ Удален файл: {os.path.basename(file_path)} (для '{route_name_to_clean}')")
                        deleted_something_for_route = True
                except OSError as e:
                    print(f"  ❌ Ошибка при удалении файла {file_path}: {e}")
            
            # Удаляем из кеша (summary и routes)
            cache_key = sanitize_filename(route_name_to_clean) # Используем то же самое имя файла
            route_popped = route_data.routes.pop(cache_key, None)
            summary_popped = route_data.summary.pop(cache_key, None)
            
            if route_popped is not None:
                 print(f"  🗑️ Удалена запись из route_data.routes для '{route_name_to_clean}'")
                 deleted_something_for_route = True
            if summary_popped is not None:
                 print(f"  🗑️ Удалена запись из route_data.summary для '{route_name_to_clean}'")
                 deleted_something_for_route = True
                 
            if deleted_something_for_route:
                routes_cleaned_count += 1

        if routes_cleaned_count > 0:
            print(f"💾 Сохраняем очищенный route_data...")
            route_data.save_to_disk() # Сохраняем изменения в кеше (удаление)
        else:
            print("ℹ️ Не найдено старых данных для очистки.")
            
        # Сохраняем ОРИГИНАЛЬНЫЕ имена маршрутов из ТЕКУЩЕГО файла
        routes_file = os.path.join(DATA_FOLDER, "original_route_names.json")
        with open(routes_file, 'w', encoding='utf-8') as f:
            json.dump({"routes": route_names}, f, ensure_ascii=False, indent=2)
        print(f"💾 Актуальный список маршрутов сохранен в {os.path.basename(routes_file)}")
        
        # --- Запуск парсинга и обработки для каждого маршрута --- 
        all_exceptions = []
        processed_route_names = [] # Сохраняем имена успешно обработанных маршрутов
        parsing_failed_routes = [] # Сохраняем имена маршрутов, где парсинг упал
        print("\n🚚 Запускаем парсинг и обработку маршрутов...")
        
        # --- ШАГ 1: Только парсинг и сбор исключений --- 
        for route_name in route_names:
            addresses_with_rows = all_routes_addresses[route_name]
            print(f"\n--- Парсинг маршрута: {route_name} ---")
            
            # 1. Запускаем парсинг через LLM
            print(f"  💬 Запуск parsing_route.py...")
            openrouter_key = os.getenv("OPENROUTER_API_KEY", "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")
            if not openrouter_key:
                print("  ❌ ОШИБКА: Ключ OpenRouter не найден (OPENROUTER_API_KEY).")
                raise ValueError("Ключ OpenRouter API не настроен.")

            parsing_cmd = [
                 sys.executable, 
                 os.path.abspath(os.path.join(BASE_DIR, "..", "parsing_route.py")),
                 "--excel", path_to_read,
                 "--openrouter_key", openrouter_key, 
                 "--model", "google/gemini-flash-1.5",
                 "--route", route_name
            ]
            cmd_str = ' '.join(parsing_cmd)
            print(f"  Выполняем команду: {cmd_str}")
            parsing_result = subprocess.run(parsing_cmd, capture_output=True, text=True, encoding='utf-8', cwd=ROOT_DIR)
            
            if parsing_result.returncode != 0:
                print(f"  ❌ Ошибка выполнения parsing_route.py для '{route_name}'")
                print(f"  Stderr:\n{parsing_result.stderr}")
                parsing_failed_routes.append(route_name) # Запоминаем, что парсинг упал
                continue # Пропускаем сбор исключений для этого маршрута
            else:
                print(f"  ✅ parsing_route.py выполнен успешно.")
                # --- ОБНОВЛЕННАЯ ЛОГИКА ПАРСИНГА ИСКЛЮЧЕНИЙ ИЗ STDOUT --- 
                lines = parsing_result.stdout.splitlines()
                route_exceptions = [] # Используем этот список для сбора исключений
                
                for line in lines:
                    # Ищем строки, соответствующие НОВОМУ формату вывода исключений
                    match = re.match(r"Маршрут\s+'(.*?)',\s+Строка\s+(\d+):\s*(.*)", line.strip(), re.IGNORECASE)
                    if match:
                        route_nm_parsed = match.group(1).strip()
                        row_num_str = match.group(2)
                        original_address_parsed = match.group(3).strip()
                        
                        # Убираем возможные кавычки по краям адреса
                        if (original_address_parsed.startswith("'") and original_address_parsed.endswith("'")) or \
                           (original_address_parsed.startswith('"') and original_address_parsed.endswith('"')):
                            original_address_parsed = original_address_parsed[1:-1]
                        
                        try:
                            row_num = int(row_num_str)
                            # Ищем оригинальный адрес по номеру строки для ЭТОГО маршрута
                            # Важно: route_name здесь - это имя маршрута из текущей итерации внешнего цикла!
                            # Нужно найти правильный addresses_with_rows для route_nm_parsed
                            current_route_addresses_rows = all_routes_addresses.get(route_nm_parsed)
                            if current_route_addresses_rows:
                                original_address_excel = next((addr for r, addr in current_route_addresses_rows if r == row_num), original_address_parsed)
                            else:
                                original_address_excel = original_address_parsed # Если не нашли список адресов для маршрута
                                
                            route_exceptions.append({
                                "row": row_num,
                                "address": original_address_excel,
                                "route": route_nm_parsed, # Используем имя из строки лога
                                "corrected": ""
                            })
                            print(f"    ✅ [Из stdout] Распознано исключение: Маршрут '{route_nm_parsed}', Строка {row_num}, Адрес: '{original_address_excel}'")
                        except ValueError:
                            print(f"    ⚠️ [Из stdout] Не удалось распознать номер строки в исключении: {line.strip()}")
                        except Exception as e_inner:
                             print(f"    ⚠️ [Из stdout] Ошибка при обработке строки исключения '{line.strip()}': {e_inner}")
                             
                # Добавляем ВСЕ найденные исключения (из route_exceptions) в общий список
                all_exceptions.extend(route_exceptions)
                # --- КОНЕЦ ОБНОВЛЕННОЙ ЛОГИКИ ПАРСИНГА --- 

        # --- ШАГ 2: Проверка наличия исключений и решение --- 
        if all_exceptions:
            print("\n❗️ Найдены адреса, требующие проверки. Отправка исключений на фронтенд.")
            # Если есть исключения, НЕ запускаем геокодирование/расчет, возвращаем исключения
            return JSONResponse({
                "status": "needs_correction", 
                "exceptions": all_exceptions,
                "routes": route_names # Отправляем все имена маршрутов, чтобы селектор обновился
            })
        else:
            print("\n✅ Адреса не требуют ручной проверки. Запускаем геокодирование и расчет...")
            # Если исключений НЕТ, запускаем геокодирование и расчет для ВСЕХ маршрутов
            # (Пропускаем те, где упал парсинг, если такие были)
            processed_route_names = [] # Список успешно обработанных
            for route_name in route_names:
                if route_name in parsing_failed_routes:
                    print(f"  ⚠️ Пропуск геокодирования/расчета для '{route_name}', т.к. парсинг не удался.")
                    continue
                
                print(f"\n--- Геокод+Расчет для: {route_name} ---")
                # --- ИЗМЕНЕНО: Передаем дату отчета из route_data --- 
                process_success = process_route(route_name, report_date_str=route_data.report_date_str) 
                if process_success:
                    print(f"  ✅ Маршрут '{route_name}' успешно обработан.")
                    processed_route_names.append(route_name)
                else:
                    print(f"  ⚠️ Маршрут '{route_name}' не был полностью обработан (ошибка в process_route).")
            
            # Очищаем CSV сводки только если были успешно обработанные маршруты
            if processed_route_names and os.path.exists(summary_csv_path):
                 try:
                     os.remove(summary_csv_path)
                     print(f"  🗑️ Удален старый файл сводки: {os.path.basename(summary_csv_path)} (т.к. есть новые данные)")
                 except OSError as e:
                     print(f"  ❌ Ошибка при удалении файла {summary_csv_path}: {e}")
                     
            print("\n🏁 Обработка всех маршрутов из файла завершена (без исключений).")
            # Возвращаем статус processed и список УСПЕШНО обработанных маршрутов
            return JSONResponse({
                "status": "processed", 
                "exceptions": [], # Пустой список исключений
                "routes": processed_route_names 
            })

    except ValueError as ve:
        print(f"❌ Ошибка при обработке файла (ValueError): {ve}")
        # import traceback # Для отладки
        # traceback.print_exc()
        return JSONResponse(status_code=400, content={"error": "Ошибка в данных Excel", "details": str(ve)})
    except subprocess.CalledProcessError as cpe:
        print(f"❌ Ошибка при вызове подпроцесса: {cpe}")
        print(f"   Stderr: {cpe.stderr}")
        return JSONResponse(status_code=500, content={"error": "Ошибка при обработке маршрута", "details": "Не удалось выполнить один из этапов обработки."}) 
    except Exception as e:
        print(f"❌ Непредвиденная ошибка при обработке файла: {e}")
        import traceback # Важно для отладки неожиданных ошибок
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": "Внутренняя ошибка сервера", "details": "Произошла ошибка при обработке файла."}) 
    
    finally:
        # --- ШАГ 3.5: Очистка временных файлов --- 
        print("--- [Finally] Очистка временных файлов --- ")
        # Удаляем исправленный файл, если он был создан
        if corrected_path and os.path.exists(corrected_path):
             try:
                 os.remove(corrected_path)
                 print(f"INFO: [Finally] Удален исправленный временный файл: {corrected_path}")
             except Exception as e_fin_corr:
                 print(f"WARNING: [Finally] Не удалось удалить исправленный временный файл {corrected_path}: {e_fin_corr}")
        
        # Всегда удаляем исходный временный файл загрузки
        if temp_upload_path and os.path.exists(temp_upload_path):
             try:
                 os.remove(temp_upload_path)
                 print(f"INFO: [Finally] Удален исходный временный файл: {temp_upload_path}")
             except Exception as e_fin_orig:
                 print(f"WARNING: [Finally] Не удалось удалить исходный временный файл {temp_upload_path}: {e_fin_orig}")
        # --- Конец Шага 3.5 --- 

@app.post("/api/submit-corrections")
async def submit_corrections(data: dict = Body(...)):
    corrections = data.get("corrections", [])
    print("📥 Получены исправления:", corrections)

    # Создаем словарь для быстрого доступа к исправлениям по *нормализованному* route_name и excel_row
    corrections_map = {}
    original_names_map = {} # Для хранения оригинального имени по нормализованному
    routes_to_process_normalized = set() # Используем нормализованные имена для обработки

    for corr in corrections:
        original_route = corr.get("route", "")
        if not original_route:
            continue # Пропускаем исправления без имени маршрута
            
        normalized_route = sanitize_filename(original_route)
        if not normalized_route:
             continue # Пропускаем, если имя стало пустым после нормализации
             
        routes_to_process_normalized.add(normalized_route)
        original_names_map[normalized_route] = original_route # Запоминаем оригинальное имя
        
        row_num_str = str(corr["row"])
        if normalized_route not in corrections_map:
            corrections_map[normalized_route] = {}
        corrections_map[normalized_route][row_num_str] = corr["corrected"]

    # --- Loop 1: Modify parsed_addresses_*.csv files --- 
    processed_routes_in_loop1 = set()
    for normalized_route in routes_to_process_normalized:
        original_name = original_names_map[normalized_route]
        parsed_path = os.path.join(PARSED_ADDRESSES_FOLDER, f"parsed_addresses_{normalized_route}.csv")
        
        if not os.path.exists(parsed_path):
            print(f"⚠️ Файл {parsed_path} не найден для маршрута '{original_name}' в цикле 1")
            continue

        try:
            with open(parsed_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                fieldnames = reader.fieldnames or ["excel_row", "route_name", "normalized_address"]

            # Вносим правки для текущего нормализованного маршрута
            route_corrections_for_file = corrections_map.get(normalized_route, {})
            modified = False
            for row in rows:
                # Нормализуем имя маршрута из строки CSV для сравнения
                row_original_name = row.get('route_name')
                if sanitize_filename(row_original_name) == normalized_route:
                    row_num_str = row.get("excel_row")
                    if row_num_str in route_corrections_for_file:
                        corrected_address = route_corrections_for_file[row_num_str]
                        if row.get('normalized_address') != corrected_address:
                             print(f"✏️ [Файл {original_name}] Обновляем строку {row_num_str}: '{row.get('normalized_address')}' -> '{corrected_address}'")
                             row['normalized_address'] = corrected_address
                             modified = True
            
            # Сохраняем, только если были изменения
            if modified:
                with open(parsed_path, "w", encoding="utf-8", newline="") as f:
                      writer = csv.DictWriter(f, fieldnames=fieldnames)
                      writer.writeheader()
                      writer.writerows(rows)
                      print(f"💾 Файл {parsed_path} обновлен исправлениями.")
                 
            processed_routes_in_loop1.add(normalized_route) # Отмечаем нормализованное имя
            
        except Exception as e:
            print(f"❌ Ошибка при обновлении файла {parsed_path}: {e}")

    # --- Loop 2: Geocode and Calculate Distance FOR ALL ROUTES --- 
    print("\n--- Запуск геокодирования и расчета расстояний для ВСЕХ маршрутов после исправлений ---")
    
    # Загружаем ПОЛНЫЙ список оригинальных имен из файла
    try:
        with open(os.path.join(DATA_FOLDER, "original_route_names.json"), 'r', encoding='utf-8') as f:
            original_names_data = json.load(f)
            all_original_route_names = original_names_data.get("routes", [])
            print(f"  Будут обработаны маршруты: {', '.join(all_original_route_names)}")
    except Exception as e:
        print(f"❌ Не удалось загрузить полный список маршрутов из original_route_names.json: {e}")
        # Если не удалось загрузить, пытаемся обработать хотя бы исправленные
        all_original_route_names = [original_names_map[norm_name] for norm_name in routes_to_process_normalized if norm_name in original_names_map]
        print(f"  ⚠️ Используется список только исправленных маршрутов: {', '.join(all_original_route_names)}")
        if not all_original_route_names:
             return JSONResponse(status_code=500, content={"error": "Ошибка обработки", "details": "Не удалось определить список маршрутов для обработки после исправлений."}) 

    # Перебираем ВСЕ оригинальные имена из загруженного файла
    for original_name in all_original_route_names:
        normalized_route = sanitize_filename(original_name)
        if not normalized_route: continue # Пропускаем если имя некорректно

        parsed_path = os.path.join(PARSED_ADDRESSES_FOLDER, f"parsed_addresses_{normalized_route}.csv")
        geocoded_path = os.path.join(GEOCODED_RESULTS_FOLDER, f"geocoded_results_{normalized_route}.csv")

        if not os.path.exists(parsed_path):
            print(f"⚠️ Файл {os.path.basename(parsed_path)} не найден для маршрута '{original_name}' в цикле 2. Пропуск.")
            continue
        
        # Вызываем process_route, который делает и геокод, и расчет, и обновление route_data
        print(f"\n--- Обработка (process_route): {original_name} ---")
        try:
            process_success = process_route(original_name)
            if process_success:
                 print(f"✅ Маршрут '{original_name}' успешно обработан через process_route.")
            else:
                 print(f"⚠️ Маршрут '{original_name}' не был полностью обработан через process_route.")
        except Exception as e_proc:
            print(f"❌ Исключение при вызове process_route для '{original_name}': {e_proc}")
            # import traceback # Раскомментировать для детальной отладки
            # traceback.print_exc()

    # Получаем обновленный список маршрутов для возврата клиенту (используем оригинальные имена)
    try:
        routes_response = get_routes() # get_routes уже возвращает оригинальные имена
        routes_list = routes_response.get("routes", [])
    except Exception as e:
        print(f"❌ Ошибка при получении списка маршрутов: {str(e)}")
        routes_list = []
        
    # Возвращаем статус об успешном сохранении
    return JSONResponse(status_code=200, content={"status": "saved", "routes": routes_list})

@app.post("/api/summary/update")
async def update_summary_endpoint(data: dict = Body(...)):
    original_route_name = data.get("route_name")
    report_distance = data.get("report_distance")
    # Получаем поля времени отчета
    report_hours = data.get("report_duration_hours")
    report_minutes = data.get("report_duration_minutes")
    # Получаем комментарий
    comment = data.get("comment")

    # !!! Убеждаемся, что service_time_per_stop_minutes НЕ принимается и НЕ передается в update_summary_item !!!
    # Это поле теперь глобальное и устанавливается через /api/upload

    # <<< Логгирование полученных данных >>>
    print(f"--- Received /api/summary/update for: {original_route_name}")
    print(f"    report_distance: '{report_distance}'")
    print(f"    report_duration_hours: '{report_hours}'")
    print(f"    report_duration_minutes: '{report_minutes}'")
    print(f"    comment: '{comment}'")
    # <<< Конец логгирования >>>

    if not original_route_name:
        raise HTTPException(status_code=400, detail="Необходимо указать имя маршрута")

    # Обновляем данные в route_data, вызывая функцию ТОЛЬКО с данными отчета
    updated = route_data.update_summary_item(
        name=original_route_name,
        report_distance=report_distance,
        report_hours=report_hours,
        report_minutes=report_minutes,
        comment=comment
        # service_time_per_stop_minutes НЕ передается
    )

    if updated:
        # CSV пока не трогаем по вашему указанию
        # csv_path = save_summary_to_csv(route_data.summary)
        # print(f"✅ Данные сохранены в CSV: {csv_path}")
        print(f"✅ Данные для '{original_route_name}' обновлены в памяти.")
    else:
        print(f"ℹ️ Данные для маршрута '{original_route_name}' не изменились.")

    # Возвращаем обновленные данные для всей сводки
    summary_response = get_summary_endpoint() # Вызываем обновленный эндпоинт
    # Добавляем статус "ok" к ответу от get_summary_endpoint
    return {"status": "ok", **summary_response}

# --- НОВЫЙ ENDPOINT ДЛЯ ЭКСПОРТА СВОДКИ --- 
@app.post("/api/export-summary")
async def export_summary_endpoint():
    print("-- Export summary requested --")
    global route_data
    if not route_data or not route_data.summary:
        print("   No summary data found for export.")
        raise HTTPException(status_code=404, detail="Нет данных для экспорта сводки.")
    try:
        # Используем метод get_summary() экземпляра route_data
        summary_data = route_data.get_summary() # Этот метод сам обновляет данные
        print(f"   Got {len(summary_data)} items for summary export.")

        if not summary_data:
             print("   get_summary() returned empty data.")
             raise HTTPException(status_code=404, detail="Нет данных для экспорта сводки (метод вернул пусто).")

        summary_values_list = list(summary_data.values())
        if not summary_values_list:
            print("   Summary dictionary is empty after getting values.")
            raise HTTPException(status_code=404, detail="Нет данных для экспорта сводки (словарь пуст).")

        df = pd.DataFrame(summary_values_list)

        # --- СОРТИРОВКА DataFrame по исходному порядку маршрутов --- 
        original_route_order = []
        try:
            with open(os.path.join(DATA_FOLDER, "original_route_names.json"), 'r', encoding='utf-8') as f:
                original_names_data = json.load(f)
                original_route_order = original_names_data.get("routes", [])
            
            if original_route_order and 'original_name' in df.columns:
                # Преобразуем колонку original_name в категориальный тип с нужным порядком
                df['original_name'] = pd.Categorical(df['original_name'], categories=original_route_order, ordered=True)
                df = df.sort_values('original_name')
                # Сбрасываем индекс после сортировки
                df = df.reset_index(drop=True)
                print(f"   DataFrame отсортирован по исходному порядку: {original_route_order}")
            else:
                print("   ⚠️ Не удалось отсортировать DataFrame по исходному порядку (список пуст или нет колонки original_name).")
        except Exception as e:
            print(f"   ⚠️ Ошибка при чтении/сортировке по original_route_names.json: {e}")
        # --- КОНЕЦ СОРТИРОВКИ --- 

        # --- ФОРМАТИРОВАНИЕ ВРЕМЕНИ ОТЧЕТА --- 
        def format_report_time(row):
            hours = row.get('report_duration_hours')
            minutes = row.get('report_duration_minutes')
            
            # Обрабатываем NaN/None как 0, но только если хотя бы одно значение не None
            has_hours = hours is not None and pd.notna(hours)
            has_minutes = minutes is not None and pd.notna(minutes)
            
            if not has_hours and not has_minutes:
                return None # Возвращаем None, если оба значения отсутствуют
                
            h = int(hours) if has_hours else 0
            m = int(minutes) if has_minutes else 0
            
            # Валидация (на всякий случай, хотя на бэке уже есть)
            if h < 0: h = 0
            if m < 0: m = 0
            if m >= 60: m = 59 
            
            return f"{h} ч {m} мин"

        # Применяем функцию для создания новой колонки
        if 'report_duration_hours' in df.columns and 'report_duration_minutes' in df.columns:
            df['report_duration_formatted'] = df.apply(format_report_time, axis=1)
            print("   Column 'report_duration_formatted' created.")
        else:
             print("   Warning: Columns 'report_duration_hours' or 'report_duration_minutes' not found in df. Skipping report time formatting.")
             df['report_duration_formatted'] = None # Создаем пустую колонку, чтобы избежать ошибок ниже
        # --- КОНЕЦ ФОРМАТИРОВАНИЯ --- 

        # --- Выбираем и переименовываем колонки --- 
        columns_to_export = {
            "original_name": "Маршрут",
            "driver_name": "ФИО водителя",
            "distance": "Расстояние (км)",
            "total_route_time_formatted": "Время на маршруте",
            "report_distance": "Расстояние (отчет)",
            "report_duration_formatted": "Время (отчет)", # <-- Используем новую колонку
            "distance_difference": "Разница (км)",
            "time_difference_formatted": "Разница (время)",
            "number_of_stops": "Кол-во точек",
            "comment": "Комментарий"
        }
        
        # Проверяем наличие колонок перед экспортом
        export_df = pd.DataFrame()
        ordered_columns = [] # Сохраняем порядок колонок
        for original_col, new_name in columns_to_export.items():
            ordered_columns.append(new_name) # Добавляем в том порядке, как они в словаре
            if original_col in df.columns:
                export_df[new_name] = df[original_col]
            else:
                print(f"   Warning: Column '{original_col}' not found in summary data for export.")
                export_df[new_name] = None # Или pd.NA или другое значение по умолчанию
        
        # Упорядочиваем колонки DataFrame
        export_df = export_df[ordered_columns]
        # --- КОНЕЦ ВЫБОРА КОЛОНОК --- 
        
        # --- Очистка ячеек разницы, если нет данных отчета --- 
        export_df_copy = export_df.copy()

        export_df_copy.loc[pd.isna(export_df_copy['Расстояние (отчет)']), 'Разница (км)'] = None

        report_time_is_empty_in_source = pd.isna(df['report_duration_hours']) & pd.isna(df['report_duration_minutes'])
        export_df_copy.loc[report_time_is_empty_in_source.values, 'Разница (время)'] = None # Используем .values для избежания проблем с индексами
        
        export_df = export_df_copy 
        # --- Конец очистки --- 

        print(f"   DataFrame prepared for export with columns: {list(export_df.columns)}")

        # 2. Создаем Excel в памяти
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Сводка')
            
            # --- СТИЛИЗАЦИЯ --- 
            workbook = writer.book
            worksheet = writer.sheets['Сводка']
            
            # Условное форматирование
            light_red_fill = PatternFill(start_color='FFFFCCCC', end_color='FFFFCCCC', fill_type='solid')
            DISTANCE_THRESHOLD = -40
            TIME_THRESHOLD_SECONDS = -5400 # -1 час 30 минут
            
            for row_idx in range(2, worksheet.max_row + 1):
                df_idx = row_idx - 2
                try:
                    dist_diff_val = df.loc[df_idx, 'distance_difference']
                    time_diff_sec_val = df.loc[df_idx, 'time_difference_seconds']
                    highlight_row = False
                    if dist_diff_val is not None and pd.notna(dist_diff_val):
                        if float(dist_diff_val) <= DISTANCE_THRESHOLD: highlight_row = True
                    if not highlight_row and time_diff_sec_val is not None and pd.notna(time_diff_sec_val):
                         if int(time_diff_sec_val) <= TIME_THRESHOLD_SECONDS: highlight_row = True
                    if highlight_row:
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = light_red_fill
                except (KeyError, IndexError, ValueError, TypeError) as fmt_err:
                     print(f"   Warning: Skipping conditional formatting for row {row_idx} due to error: {fmt_err}")
            print(f"   Conditional formatting applied.")

            # Границы
            thin_black_border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'), top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'))
            thin_gray_border = Border(left=Side(style='thin', color='FFA9A9A9'), right=Side(style='thin', color='FFA9A9A9'), top=Side(style='thin', color='FFA9A9A9'), bottom=Side(style='thin', color='FFA9A9A9'))
            for cell in worksheet["1:1"]: cell.border = thin_black_border
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column): 
                for cell in row: cell.border = thin_gray_border
            print(f"   Borders applied.")

            # Ширина и перенос
            column_widths_px = {
                'Маршрут': 380, 'ФИО водителя': 300, 'Расстояние (км)': 140,
                'Время на маршруте': 140, 'Расстояние (отчет)': 140, 'Время (отчет)': 140,
                'Разница (км)': 120, 'Разница (время)': 120, 'Кол-во точек': 120,
                'Комментарий': 380
            }
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            for i, column_name in enumerate(export_df.columns):
                column_letter = get_column_letter(i + 1)
                width_px = column_widths_px.get(column_name, 100)
                excel_width = width_px / 9.0 
                worksheet.column_dimensions[column_letter].width = excel_width
            for row in worksheet.iter_rows(): 
                for cell in row: cell.alignment = wrap_alignment
            print(f"   Widths and wrap text applied.")
            # --- КОНЕЦ СТИЛИЗАЦИИ --- 

        buffer.seek(0)
        print("   Excel buffer created.")

        # 3. Готовим имя файла и возвращаем ответ
        # --- ЛОГИКА ФОРМИРОВАНИЯ ИМЕНИ ФАЙЛА (ОСТАВЛЯЕМ НОВУЮ) --- 
        filename = "summary_nodata.xlsx" # Имя по умолчанию
        report_date_str = route_data.report_date_str # Получаем сохраненную дату
        print(f"   Report date string from route_data: '{report_date_str}'")
        if report_date_str:
            try:
                report_date_obj = datetime.strptime(report_date_str, "%Y-%m-%d") 
                formatted_date = report_date_obj.strftime("%d.%m.%Y")
                filename = f"summary_{formatted_date}.xlsx"
                print(f"   Generated filename with date: {filename}")
            except ValueError as e:
                print(f"   ⚠️ Error parsing report date '{report_date_str}': {e}. Using default filename.")
                filename = "summary_nodata.xlsx"
        else:
             print("   Report date string is empty. Using default filename.")
        # --- КОНЕЦ ЛОГИКИ ФОРМИРОВАНИЯ ИМЕНИ ФАЙЛА ---

        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"' # Используем новое имя
        }
        
        print(f"   Returning file: {filename}")
        return StreamingResponse(
            buffer,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )

    except Exception as e:
        import traceback
        print(f"!!! Error during summary export: {e}")
        print(traceback.format_exc()) # Печатаем полный traceback для диагностики
        raise HTTPException(status_code=500, detail=f"Ошибка при генерации Excel файла: {e}")

# --- КОНЕЦ ENDPOINT'А ЭКСПОРТА ---

# --- НОВЫЙ ЭНДПОИНТ ДЛЯ ОБНОВЛЕНИЯ ВРЕМЕНИ НА ТОЧКУ ---
class ServiceTimeUpdate(BaseModel):
    service_time: int

@app.post("/api/update-service-time")
async def update_service_time_endpoint(data: ServiceTimeUpdate):
    print(f"--- Received /api/update-service-time with: {data.service_time} min ---")
    try:
        # 1. Устанавливаем новое глобальное время (только сохраняем)
        time_updated = route_data.set_global_service_time(data.service_time)

        # 2. Запускаем пересчет сводки и получаем обновленные данные
        # Метод get_summary теперь отвечает за актуализацию и пересчет
        updated_summary_data = route_data.get_summary()
        
        # 3. Формируем ответ, аналогичный /api/summary
        result = []
        for norm_name, item_data in updated_summary_data.items():
            result.append({
                "route_name": item_data.get("original_name", norm_name),
                "driver_name": item_data.get("driver_name", "—"),
                "distance": item_data.get("distance", "Н/Д"),
                "duration_seconds": item_data.get("duration_seconds"),
                "duration": item_data.get("duration_formatted", "Н/Д"),
                "report_distance": item_data.get("report_distance"),
                "report_duration_hours": item_data.get("report_duration_hours"),
                "report_duration_minutes": item_data.get("report_duration_minutes"),
                "distance_difference": item_data.get("distance_difference"),
                "time_difference_formatted": item_data.get("time_difference_formatted"),
                "time_difference_seconds": item_data.get("time_difference_seconds"),
                "total_route_time_seconds": item_data.get("total_route_time_seconds"),
                "total_route_time_formatted": item_data.get("total_route_time_formatted", "Н/Д"),
                "number_of_stops": item_data.get("number_of_stops", "Н/Д")
            })
        
        return {
            "status": "ok",
            "summary": result,
            "global_service_time_minutes": route_data.global_service_time_minutes
        }

    except Exception as e:
        print(f"❌ Error in /api/update-service-time: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка при обновлении времени на точку: {e}")
# --- КОНЕЦ НОВОГО ЭНДПОИНТА ---

# --- НОВЫЙ ENDPOINT ДЛЯ ПОЛУЧЕНИЯ ВСЕХ ДАННЫХ МАРШРУТОВ --- 
@app.get("/api/all-route-data")
def get_all_route_data_endpoint():
    print("--- Request received for /api/all-route-data ---")
    global route_data
    all_routes_data_for_frontend = {}
    route_names = []

    # 1. Получаем список актуальных оригинальных имен маршрутов
    try:
        with open(os.path.join(DATA_FOLDER, "original_route_names.json"), 'r', encoding='utf-8') as f:
            original_names_data = json.load(f)
            route_names = original_names_data.get("routes", [])
        if not route_names:
             print("  original_route_names.json пуст или не содержит маршрутов.")
             return {}
        print(f"  Будут загружены данные для маршрутов: {route_names}")
    except Exception as e:
        print(f"  ❌ Ошибка чтения original_route_names.json: {e}")
        raise HTTPException(status_code=500, detail="Не удалось получить список маршрутов.")

    # 2. Итерируем по списку и собираем данные для каждого
    for original_name in route_names:
        sanitized_name = sanitize_filename(original_name)
        if not sanitized_name: continue

        print(f"  -> Обработка данных для: {original_name}")
        try:
            # Используем существующую логику get_route_data_endpoint,
            # но без HTTP исключений, чтобы одна ошибка не сломала все
            single_route_data = get_route_data_endpoint(original_name)
            
            # Проверяем на наличие ошибки внутри ответа от get_route_data_endpoint
            if isinstance(single_route_data, dict) and not single_route_data.get("error"): 
                # Используем оригинальное имя как ключ для фронтенда
                all_routes_data_for_frontend[original_name] = single_route_data
                print(f"     ✅ Данные для '{original_name}' успешно получены.")
            else:
                error_msg = single_route_data.get("message", "Неизвестная ошибка") if isinstance(single_route_data, dict) else "Некорректный формат ответа"
                print(f"     ⚠️ Не удалось получить данные для '{original_name}': {error_msg}")
                # Можно решить, добавлять ли пустую запись или пропускать
                # all_routes_data_for_frontend[original_name] = {"name": original_name, "error": True, "message": error_msg}
        except Exception as e_single:
            print(f"     ❌ Исключение при получении данных для '{original_name}': {e_single}")
            # Можно решить, добавлять ли пустую запись или пропускать
            # all_routes_data_for_frontend[original_name] = {"name": original_name, "error": True, "message": str(e_single)}
            
    print(f"--- /api/all-route-data: Возвращаем данные для {len(all_routes_data_for_frontend)} маршрутов --- ")
    return JSONResponse(content=all_routes_data_for_frontend)
# --- КОНЕЦ НОВОГО ENDPOINT'А --- 

# --- ДОБАВЛЕНО: Модель для валидации входных данных пересчета ---
class RecalculatePoint(BaseModel):
    originalIndex: int
    address: str
    isOffice: bool
    isHidden: bool # Хотя клиент не должен их слать, примем на всякий случай
    isModified: bool
    lat: Optional[float] = None
    lon: Optional[float] = None

class RecalculateRequest(BaseModel):
    routeId: str # Используем routeId, которое клиент присылает
    routeName: str # Имя маршрута, как оно есть
    points: List[RecalculatePoint]

# --- ДОБАВЛЕН НОВЫЙ ЭНДПОИНТ ДЛЯ ПЕРЕСЧЕТА ---
@app.post("/api/recalculate-route")
# --- ИЗМЕНЕНО: Вернули data: RecalculateRequest --- 
async def recalculate_route_endpoint(data: RecalculateRequest):
    """
    Обрабатывает запрос на пересчет маршрута с измененными/добавленными/удаленными/скрытыми точками.
    """
    # --- ИЗМЕНЕНО: Раскомментировано начало логики --- 
    print(f"--- Received recalculate request for routeId: {data.routeId} ---")
    # Получаем routeName из запроса, т.к. routeId может быть некорректным ключом
    route_name = data.routeName 
    sanitized_route_name = sanitize_filename(route_name)
    if not sanitized_route_name:
         raise HTTPException(status_code=400, detail="Invalid route name provided.")
    
    # --- Определяем пути к файлам --- 
    # --- ИЗМЕНЕНО: Исправлен путь к файлу геоданных ---
    geocoded_file_path = os.path.join(config.GEOCODED_DIR, f"{sanitized_route_name}_geocoded.json") # Убран префикс geocoded_results_
    # --- КОНЕЦ ИЗМЕНЕНИЯ ---
    # --- ИЗМЕНЕНО: Исправлено имя файла для distance_data ---
    distance_data_file_path = os.path.join(config.ROUTE_RESULTS_DIR, f"{sanitized_route_name}_distance_data.json") # Имя как в get_route_data
    # --- КОНЕЦ ИЗМЕНЕНИЯ ---
    # --- ДОБАВЛЕНО: Путь к файлу с результатами для загрузки start_time --- 
    route_results_file_path = os.path.join(config.ROUTE_RESULTS_DIR, f"route_results_{sanitized_route_name}.json")
    print(f"[Recalc] Paths: geocoded={geocoded_file_path}, distance={distance_data_file_path}, results={route_results_file_path}")
    # --- КОНЕЦ ДОБАВЛЕНИЯ ---

    # --- Загружаем ТЕКУЩИЕ данные --- 
    current_geocoder_output = []
    if os.path.exists(geocoded_file_path):
        try:
            with open(geocoded_file_path, 'r', encoding='utf-8') as f:
                current_geocoder_output = json.load(f)
        except Exception as e_load_geo:
            print(f"ERROR loading current geocoded data {geocoded_file_path}: {e_load_geo}")
            # Возможно, стоит вернуть ошибку, если не можем загрузить основу
            raise HTTPException(status_code=500, detail=f"Не удалось загрузить текущие геоданные для маршрута {route_name}")
    else:
         print(f"ERROR: Geocoded file not found for recalculation: {geocoded_file_path}")
         raise HTTPException(status_code=404, detail=f"Геоданные для маршрута {route_name} не найдены.")

    # --- ДОБАВЛЕНО: Загружаем start_time_used из файла результатов --- 
    original_start_time_iso = None
    if os.path.exists(route_results_file_path):
        try:
            with open(route_results_file_path, 'r', encoding='utf-8') as f:
                route_results_data = json.load(f)
                original_start_time_iso = route_results_data.get("start_time_used")
                if original_start_time_iso:
                    print(f"[Recalc] Loaded original start_time_used: {original_start_time_iso} from {route_results_file_path}")
                else:
                    print(f"[Recalc] WARNING: 'start_time_used' not found in {route_results_file_path}. Recalculation might use default time.")
                    # Альтернатива: вызвать get_start_time_iso() без даты? Или вернуть ошибку?
                    # Пока оставим None, get_route_segments должен будет обработать это
        except Exception as e_load_results:
            print(f"[Recalc] ERROR loading route results file {route_results_file_path}: {e_load_results}. Proceeding without original start time.")
    else:
        print(f"[Recalc] WARNING: Route results file not found: {route_results_file_path}. Cannot load original start_time_used.")
        # Обработка случая, если файла нет (первый расчет еще не завершился?)
        # Можно попробовать вычислить время по умолчанию, но лучше требовать наличие файла.
        # Пока оставим original_start_time_iso = None
    
    # --- ВАЖНО: Проверяем, что время удалось загрузить --- 
    if not original_start_time_iso:
        print("CRITICAL ERROR: Cannot proceed with recalculation without original start_time_iso.")
        # Отправляем ошибку клиенту, так как без времени расчет некорректен
        raise HTTPException(status_code=500, detail=f"Не удалось загрузить исходное время расчета (start_time_used) для маршрута {route_name}. Пересчет невозможен.")
    # --- КОНЕЦ ДОБАВЛЕНИЯ ---

    # --- Обработка точек из запроса --- 
    new_geocoder_output = []
    points_for_route_calculation = []
    modified_indices = set()
    
    print(f"Processing {len(data.points)} points received from client...")
    
    # Добавляем глобальное время на обслуживание (сейчас не используется в расчете сегментов)
    # global_service_time = route_data.global_service_time_minutes * 60 # в секундах
    
    point_counter = 0
    for point_data in data.points:
        point_counter += 1
        
        if point_data.isHidden:
            print(f"  Skipping hidden point at originalIndex: {point_data.originalIndex}")
            continue
    
        geocoded_point = {
            "input": point_data.address,
            # Определяем СТАРТ/ФИНИШ более надежно
            "excel_row": "СТАРТ" if point_data.isOffice and point_counter == 1 else ("ФИНИШ" if point_data.isOffice and point_counter == len(data.points) else ""), 
            "found": None,
            "lat": point_data.lat,
            "lon": point_data.lon,
            "type": None,
            "description": None,
            # Добавляем originalIndex в результат для возможного сопоставления на клиенте
            "original_index_from_request": point_data.originalIndex 
        }
    
        needs_geocoding = point_data.isModified or (not point_data.isOffice and (point_data.lat is None or point_data.lon is None))
        
        # --- ДОБАВЛЕНО: Принудительная установка координат для офиса ---
        if point_data.isOffice:
            try:
                office_loc = config.OFFICE_LOCATION
                if office_loc and "lat" in office_loc and "lon" in office_loc:
                    geocoded_point['lat'] = office_loc['lat']
                    geocoded_point['lon'] = office_loc['lon']
                    geocoded_point['found'] = True # Офис всегда найден
                    geocoded_point['type'] = 'office'
                    geocoded_point['description'] = office_loc.get("name", "Офис РТК")
                    print(f"  Point {point_counter} ('{point_data.address}') is Office. Set coords from config: lat={geocoded_point['lat']}, lon={geocoded_point['lon']}")
                else:
                     print(f"  WARNING: Office point {point_counter} ('{point_data.address}') - OFFICE_LOCATION not configured correctly.")
                     geocoded_point['lat'] = None
                     geocoded_point['lon'] = None
                     geocoded_point['found'] = False
            except Exception as e_office:
                print(f"  ERROR setting office coordinates for point {point_counter}: {e_office}")
                geocoded_point['lat'] = None
                geocoded_point['lon'] = None
                geocoded_point['found'] = False
        # --- КОНЕЦ ДОБАВЛЕНИЯ ---
        
        if needs_geocoding:
            print(f"  Point {point_counter} ('{point_data.address}') needs geocoding (isModified: {point_data.isModified}, isOffice: {point_data.isOffice}, lat: {point_data.lat}).")
            modified_indices.add(point_data.originalIndex)
            try:
                geocode_result = geocode_address(point_data.address)
                if geocode_result and geocode_result.get('lat') is not None and geocode_result.get('lon') is not None:
                    geocoded_point['lat'] = geocode_result['lat']
                    geocoded_point['lon'] = geocode_result['lon']
                    geocoded_point['type'] = geocode_result.get('type')
                    geocoded_point['description'] = geocode_result.get('description')
                    geocoded_point['found'] = geocode_result.get('found') # Используем found из результата геокодера
                    print(f"    Geocoded successfully: lat={geocoded_point['lat']}, lon={geocoded_point['lon']}, type={geocoded_point['type']}")
                else:
                    print(f"    WARNING: Geocoding failed or returned incomplete data for address: {point_data.address}")
                    geocoded_point['lat'] = None
                    geocoded_point['lon'] = None
                    geocoded_point['description'] = "⚠️ Ошибка геокодирования"
                    geocoded_point['found'] = False # Явно указываем, что не найдено
            except Exception as e:
                print(f"    ERROR during geocoding for address '{point_data.address}': {e}")
                geocoded_point['lat'] = None
                geocoded_point['lon'] = None
                geocoded_point['description'] = "⚠️ Ошибка геокодирования"
                geocoded_point['found'] = False # Явно указываем, что не найдено
        else:
             # Если геокодирование не нужно, пытаемся найти тип/описание в СТАРЫХ данных по адресу
             # Это менее надежно, чем по индексу, но у нас нет исходного индекса в current_geocoder_output из CSV
             # Пока просто оставляем поля как есть (lat/lon из запроса)
             original_point_from_loaded_data = next((p for p in current_geocoder_output if p.get('input') == point_data.address), None)
             if original_point_from_loaded_data:
                 geocoded_point['found'] = original_point_from_loaded_data.get('found') # Может быть True/False/None
                 geocoded_point['type'] = original_point_from_loaded_data.get('type')
                 geocoded_point['description'] = original_point_from_loaded_data.get('description')
             else:
                 print(f"  Point {point_counter} ('{point_data.address}') - geocoding not needed. Using provided coords. Could not find original data by address.")
                 # Если не офис, но нет координат - ставим ошибку
                 if not point_data.isOffice and (geocoded_point['lat'] is None or geocoded_point['lon'] is None):
                      geocoded_point['description'] = "⚠️ Нет координат"
                      geocoded_point['found'] = False
                 elif point_data.isOffice:
                       geocoded_point['found'] = True # Считаем офис найденным
    
        new_geocoder_output.append(geocoded_point)
    
        # Добавляем точку для расчета маршрута ТОЛЬКО если есть координаты
        if geocoded_point.get('lat') is not None and geocoded_point.get('lon') is not None:
            points_for_route_calculation.append({"lat": geocoded_point['lat'], "lon": geocoded_point['lon']})
        else:
            # Если точка не офис и у нее нет координат - это ошибка, нельзя считать маршрут
            if not point_data.isOffice:
                print(f"ERROR: Point '{point_data.address}' has no coordinates after processing. Route calculation cannot proceed.")
                raise HTTPException(status_code=400, detail=f"Точка '{point_data.address}' не имеет координат после обработки. Невозможно пересчитать маршрут.")
    # --- КОНЕЦ БЛОКА ОБРАБОТКИ ТОЧЕК ---
    
    # --- Добавлен Print для проверки --- 
    print(f">>> Step 3 OK: Processed {len(new_geocoder_output)} points. Points for calculation: {len(points_for_route_calculation)}.")
    
    # --- УДАЛЕНО: Тестовый ответ --- 
    # return JSONResponse(content={"status": "ok", "message": "Step 3 reached!"})
    
    # --- РАСКОММЕНТИРОВАНО: Расчет маршрута и сохранение --- 
    
    # --- Пересчитываем маршрут ---
    print(f"Recalculating route with {len(points_for_route_calculation)} valid points...")
    new_distance_data = None
    # Проверяем, что точек больше одной (не только офис старт/финиш)
    if len(points_for_route_calculation) > 1:
        try:
            # Вызываем асинхронный расчет сегментов
            # TODO: Получать traffic_mode из config?
            # segments = await route_distance.get_route_segments_async(points_for_route_calculation, traffic_mode='auto') 
            
            # --- ИЗМЕНЕНО: Вызываем новую синхронную функцию С ПАРАМЕТРАМИ --- 
            print(f"[Recalc] Calling get_route_segments with {len(points_for_route_calculation)} points, start_time: {original_start_time_iso}")
            segments = route_distance.get_route_segments(points_for_route_calculation, start_time_iso=original_start_time_iso, traffic_mode='statistics')
            # --- КОНЕЦ ИЗМЕНЕНИЯ --- 
            
            # Форматируем результат с помощью вспомогательной функции
            new_distance_data = format_distance_data_from_segments(segments)
            # --- [DEBUG recalc] Лог 1: Рассчитанное время в пути --- 
            print(f"--- [DEBUG recalc Log 1] Route: {route_name}, Recalculated total_duration: {new_distance_data.get('total_duration')}")
            # --- КОНЕЦ ЛОГА 1 ---
            print(f"Route recalculated. Distance: {new_distance_data.get('formatted_distance')}, Duration: {new_distance_data.get('formatted_duration')}")
            
        except Exception as e:
            print(f"ERROR during route recalculation: {e}")
            # В случае ошибки расчета маршрута, вернем ошибку клиенту
            raise HTTPException(status_code=500, detail=f"Ошибка при пересчете маршрута: {e}")
    else:
        print("Skipping route calculation: Not enough valid points.")
        # Если точек мало, создаем пустые данные о расстоянии
        new_distance_data = {
            "segments": [],
            "total_distance": 0,
            "total_duration": 0,
            "formatted_distance": "0 км",
            "formatted_duration": "0 мин"
        }

    # --- Сохраняем обновленные данные --- 
    try:
        # Сохраняем НОВЫЙ geocoder_output (перезаписываем старый JSON или CSV)
        # --- ЛОГ 8: Перед сохранением new_geocoder_output --- 
        print(f"[DEBUG recalc] Before saving new_geocoder_output: length={len(new_geocoder_output)}")
        if len(new_geocoder_output) >= 1: print(f"  First point: {new_geocoder_output[0].get('excel_row')}, Input: {new_geocoder_output[0].get('input')}")
        if len(new_geocoder_output) >= 2: print(f"  Last point: {new_geocoder_output[-1].get('excel_row')}, Input: {new_geocoder_output[-1].get('input')}")
        # --- КОНЕЦ ЛОГА 8 ---
        with open(geocoded_file_path, 'w', encoding='utf-8') as f:
            json.dump(new_geocoder_output, f, ensure_ascii=False, indent=4)
        print(f"Saved updated geocoded data to: {geocoded_file_path}")
        # --- ДОБАВЛЕНО: Сохраняем НОВЫЕ distance_data --- 
        if new_distance_data is not None:
            # --- [DEBUG recalc] Лог 2a: Путь сохранения distance data --- 
            print(f"--- [DEBUG recalc Log 2a] Route: {route_name}, Saving new distance data to: {distance_data_file_path}")
            # --- КОНЕЦ ЛОГА 2a ---
            with open(distance_data_file_path, 'w', encoding='utf-8') as f:
                json.dump(new_distance_data, f, ensure_ascii=False, indent=4)
            # --- [DEBUG recalc] Лог 2b: Проверка существования файла ПОСЛЕ сохранения --- 
            print(f"--- [DEBUG recalc Log 2b] Route: {route_name}, File exists after save? ({os.path.basename(distance_data_file_path)}): {os.path.exists(distance_data_file_path)}")
            # --- КОНЕЦ ЛОГА 2b ---
        else:
            print(f"--- [DEBUG recalc] Route: {route_name}, Skipping distance data save (new_distance_data is None).")
        # --- КОНЕЦ ДОБАВЛЕНИЯ --- 
                 
    except Exception as e_save:
        print(f"ERROR saving updated route data: {e_save}")
        # Не прерываем выполнение, но логируем ошибку

    # --- Обновляем данные в глобальном RouteData и summary --- 
    # Собираем данные в формате, как их хранит RouteData.add_route
    
    # --- ИЗМЕНЕНО: Правильный расчет number_of_stops для сохранения --- 
    visible_in_recalc = sum(1 for p in new_geocoder_output if not p.get('hidden', False))
    stops_for_summary_recalc = max(0, visible_in_recalc - 2)
    # --- ЛОГ 9: Рассчитанное значение для сохранения --- 
    print(f"[DEBUG recalc] Calculated stops for saving: visible={visible_in_recalc}, stops_for_summary_recalc={stops_for_summary_recalc}")
    # --- КОНЕЦ ЛОГА 9 ---
    # --- [DEBUG recalc] Лог 10: Данные перед добавлением в RouteData --- 
    print(f"--- [DEBUG recalc Log 10] Route: {route_name}, Calling add_route with: number_of_stops={stops_for_summary_recalc}, total_duration={new_distance_data.get('total_duration')}")
    # --- КОНЕЦ ЛОГА 10 --- 
    
    updated_route_full_data = {
        "route_name": route_name, # Используем оригинальное имя
        "geocoder_output": new_geocoder_output,
        "route_points": points_for_route_calculation, # Передаем список lat/lon
        "distance_data": new_distance_data,
        # --- ИЗМЕНЕНО: Используем правильное значение --- 
        "number_of_stops": stops_for_summary_recalc,
        # --- КОНЕЦ ИЗМЕНЕНИЯ --- 
        "total_route_time_formatted": new_distance_data.get("formatted_duration", "Н/Д"), # Для get_route_data_endpoint
        "global_service_time_minutes": route_data.global_service_time_minutes # Для get_route_data_endpoint
    }
    # Обновляем данные в RouteData (это также пересчитает summary и сохранит)
    # Используем оригинальное имя маршрута для add_route
    route_data.add_route(route_name, updated_route_full_data)
    # Сохраняем сразу после обновления (add_route не сохраняет сам)
    route_data.save_to_disk()
    print(f"Updated route data in memory and saved for: {route_name}")
    
    # --- ДОБАВЛЕНО: Обновляем общую сводку ---
    try:
        print("Updating overall summary after recalculation...")
        route_data.get_summary() # Этот вызов пересчитает и сохранит route_summary.json
        print("Overall summary updated.")
    except Exception as e_summary_update:
        # Логируем ошибку, но не прерываем отправку ответа клиенту
        print(f"ERROR updating overall summary after recalculation: {e_summary_update}")
    # --- КОНЕЦ ДОБАВЛЕНИЯ ---

    # --- Формируем финальный ответ для клиента --- 
    # Используем только что рассчитанные/сохраненные данные
    # --- [DEBUG recalc] Лог 11: Данные из summary перед отправкой ответа --- 
    final_stops = route_data.summary.get(sanitized_route_name, {}).get("number_of_stops", "N/A")
    final_time_formatted = route_data.summary.get(sanitized_route_name, {}).get("total_route_time_formatted", "N/A")
    print(f"--- [DEBUG recalc Log 11] Route: {route_name}, Final summary data: number_of_stops={final_stops}, total_route_time_formatted='{final_time_formatted}'")
    # --- КОНЕЦ ЛОГА 11 ---
    final_response = {
        "status": "recalculated",
        "route_name": route_name,
        "geocoder_output": new_geocoder_output, # Передаем обновленный geocoder output
        "route_points": points_for_route_calculation, # Передаем точки для карты
        "distance_data": new_distance_data, # Передаем новые данные о расстоянии
        # --- ИЗМЕНЕНО: Берем актуальное значение из route_data.summary --- 
        # (Оно должно было обновиться при вызове route_data.add_route)
        "number_of_stops": route_data.summary.get(sanitized_route_name, {}).get("number_of_stops", 0),
        # --- КОНЕЦ ИЗМЕНЕНИЯ --- 
        "total_route_time_formatted": route_data.summary.get(sanitized_route_name, {}).get("total_route_time_formatted", "Н/Д"),
        "global_service_time_minutes": route_data.global_service_time_minutes
    }
    
    return JSONResponse(content=final_response)
    
    # --- КОНЕЦ ЗАКОММЕНТИРОВАННОГО БЛОКА --- 

# --- НОВАЯ ФУНКЦИЯ ГЕОКОДИРОВАНИЯ --- 
def run_geocoding(parsed_addresses_file: str, geocoded_json_path: str, original_route_name: str) -> bool:
    """Выполняет геокодирование адресов из CSV и сохраняет результат в JSON.

    Args:
        parsed_addresses_file: Путь к CSV файлу с адресами (из parsing_route.py).
        geocoded_json_path: Путь к JSON файлу для сохранения результатов геокодирования.
        original_route_name: Оригинальное имя маршрута для логов.

    Returns:
        True, если геокодирование успешно, False в случае ошибки.
    """
    print(f"--- [run_geocoding] Запуск для '{original_route_name}' ---")
    print(f"    Входной CSV: {parsed_addresses_file}")
    print(f"    Выходной JSON: {geocoded_json_path}")
    
    if not os.path.exists(parsed_addresses_file):
        print(f"  ❌ Ошибка: Файл с адресами не найден: {parsed_addresses_file}")
        return False
        
    geocoded_results = []
    success_count = 0
    error_count = 0

    try:
        with open(parsed_addresses_file, "r", encoding="utf-8") as infile:
            reader = csv.DictReader(infile)
            if not reader.fieldnames or "normalized_address" not in reader.fieldnames:
                 print(f"  ❌ Ошибка: Отсутствует колонка 'normalized_address' в файле {parsed_addresses_file}")
                 return False
                 
            for row_num, row in enumerate(reader, start=1): # Нумеруем строки для логов
                address = row.get("normalized_address")
                excel_row = row.get("excel_row", "?") # Получаем номер строки Excel
                
                if not address:
                    print(f"  ⚠️ Пропуск строки {row_num} (Excel: {excel_row}): Пустой адрес.")
                    continue
                
                try:
                    # Вызываем функцию геокодирования из geocoder.py
                    result = geocode_address(address)
                    
                    # Добавляем доп. поля, которые могут понадобиться дальше
                    result["excel_row"] = excel_row
                    result["route_name"] = original_route_name # Сохраняем оригинальное имя
                    
                    geocoded_results.append(result)
                    
                    if result.get("error"):
                        error_count += 1
                        print(f"    -> Строка {row_num} (Excel: {excel_row}): Ошибка геокодирования - {result['error']}")
                    else:
                        success_count += 1
                        # Укороченный лог успеха
                        # print(f"    -> Строка {row_num} (Excel: {excel_row}): OK - {result.get('lat')}, {result.get('lon')}")
                        
                except Exception as e_single:
                    print(f"  ❌ Критическая ошибка при геокодировании адреса '{address}' (строка {row_num}): {e_single}")
                    error_count += 1
                    # Можно добавить запись об ошибке в geocoded_results, если нужно
                    geocoded_results.append({
                        "excel_row": excel_row,
                        "route_name": original_route_name,
                        "input": address,
                        "found": None,
                        "type": "error", 
                        "description": "<span class=\"accuracy-tag tag-error\">⚠️ Внутренняя ошибка геокодера</span>", 
                        "lat": None,
                        "lon": None,
                        "error": f"Внутренняя ошибка: {e_single}"
                    })

    except Exception as e_read:
        print(f"  ❌ Ошибка чтения файла {parsed_addresses_file}: {e_read}")
        return False

    # Сохраняем результат в JSON
    try:
        # Создаем директорию, если не существует
        os.makedirs(os.path.dirname(geocoded_json_path), exist_ok=True)
        
        with open(geocoded_json_path, "w", encoding="utf-8") as outfile:
            json.dump(geocoded_results, outfile, ensure_ascii=False, indent=4)
        print(f"  ✅ Результаты геокодирования сохранены в: {geocoded_json_path}")
        print(f"     Успешно: {success_count}, Ошибок: {error_count}")
        return True # Успех, даже если были ошибки геокодирования отдельных адресов
        
    except Exception as e_save:
        print(f"  ❌ Ошибка сохранения JSON файла {geocoded_json_path}: {e_save}")
        return False
# --- КОНЕЦ НОВОЙ ФУНКЦИИ ---

# --- КОНЕЦ КЛАССА RouteData И ДРУГИХ ОПРЕДЕЛЕНИЙ ПЕРЕД ГЛОБАЛЬНЫМИ ЭНДПОИНТАМИ ---

# Глобальный экземпляр данных, если он еще не определен где-то выше
# (если он уже есть, эта строка не нужна, но для примера оставим,
# предполагая, что route_data_store должен быть доступен глобально)
# route_data_store = RouteData() # Убедитесь, что это согласуется с вашей логикой

# ... здесь идут ваши @app.get, @app.post декораторы и функции ...

# Добавляем этот блок в самый конец файла:
if __name__ == "__main__":
    import uvicorn
    # Этот код находится в frontend/main.py, и переменная FastAPI называется 'app'.
    # Uvicorn будет искать объект 'app' в текущем модуле ('main').
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=False, workers=1)

