from fastapi import FastAPI, UploadFile, Request, Form, Body, Path, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import uvicorn
import os
import shutil
import subprocess
import pandas as pd
import csv
import sys
import json
from typing import List, Dict, Any, Optional
import io
import math
import re
import numpy as np
import logging
from datetime import datetime
import config # <--- ДОБАВИТЬ, если отсутствует
# --- Добавляем импорты для стилизации Excel ---
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side # <<< ДОБАВЛЕНЫ Border, Side
# --- Конец импортов для стилизации ---

# Добавляем путь к корню проекта
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

# Импортируем напрямую из файла в корне проекта
from geocoder import geocode_address
# --- Импортируем функцию расчета расстояний --- 
import route_distance 

app = FastAPI()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.abspath(os.path.join(BASE_DIR, ".."))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploaded")
DATA_FOLDER = os.path.join(BASE_DIR, "data")
PROJECT_DATA_FOLDER = os.path.join(ROOT_DIR, "data")
GEOCODED_RESULTS_FOLDER = os.path.join(PROJECT_DATA_FOLDER, "geocoded_results")
ROUTE_RESULTS_FOLDER = os.path.join(PROJECT_DATA_FOLDER, "route_results")
PARSED_ADDRESSES_FOLDER = os.path.join(PROJECT_DATA_FOLDER, "parsed_addresses")
SUMMARY_CSV_FOLDER = os.path.join(PROJECT_DATA_FOLDER, "summary_csv")

# Создаем все необходимые директории
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DATA_FOLDER, exist_ok=True)
os.makedirs(GEOCODED_RESULTS_FOLDER, exist_ok=True)
os.makedirs(ROUTE_RESULTS_FOLDER, exist_ok=True)
os.makedirs(PARSED_ADDRESSES_FOLDER, exist_ok=True)
os.makedirs(SUMMARY_CSV_FOLDER, exist_ok=True)

app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

# Вспомогательная функция для создания безопасных имен файлов
def sanitize_filename(filename):
    """Преобразует строку в безопасное имя файла"""
    if not isinstance(filename, str):
        return "unnamed"
    # Заменяем пробелы на подчеркивания, сохраняя все остальные символы
    s = filename.strip().replace(' ', '_')
    return s if s else "unnamed"

# Вспомогательная функция для очистки данных перед JSON сериализацией
def sanitize_data_for_json(data):
    if isinstance(data, dict):
        return {k: sanitize_data_for_json(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_data_for_json(item) for item in data]
    elif isinstance(data, float):
        if math.isnan(data) or math.isinf(data):
            return None # Заменяем NaN/inf на None
        return data
    else:
        return data

# Структура для хранения обработанных данных
class RouteData:
    def __init__(self):
        self.routes = {}  # Ключ: sanitized_name
        self.summary = {} # Ключ: sanitized_name, Значение: {...}
        self.drivers = {} # <--- НОВОЕ ПОЛЕ: Словарь для хранения водителей {sanitized_name: driver_name}
        self.current_file = ""  # Имя текущего файла
        self.global_service_time_minutes = 0 # <--- НОВОЕ ГЛОБАЛЬНОЕ ПОЛЕ (по умолчанию 0)

    def add_route(self, name, data):
        sanitized_name = sanitize_filename(name)
        if not sanitized_name: return

        # Сохраняем полные данные маршрута (geocoder_output, route_points, distance_data)
        self.routes[sanitized_name] = data

        distance_data = data.get("distance_data", {})
        distance_km = distance_data.get("total_distance") # Ожидаем округленное до целых км
        duration_sec = distance_data.get("total_duration") # Ожидаем секунды
        duration_formatted = distance_data.get("formatted_duration", "Н/Д")
        # Получаем количество промежуточных точек (все точки из geocoded_file)
        num_intermediate_stops = len(data.get("route_points", []))

        # Инициализация или обновление записи в summary
        if sanitized_name not in self.summary:
            self.summary[sanitized_name] = {
                "original_name": name,
                "distance": distance_km if distance_km is not None else "Н/Д",
                "duration_seconds": duration_sec, # Храним секунды для расчета разницы и времени на маршруте
                "duration_formatted": duration_formatted, # Время в пути
                "number_of_stops": num_intermediate_stops, # <--- НОВОЕ ПОЛЕ: кол-во точек
                "report_distance": None,
                "report_duration_hours": None,
                "report_duration_minutes": None,
                "distance_difference": None,
                "time_difference_formatted": None,
                "total_route_time_seconds": None, # <--- НОВОЕ ПОЛЕ: общее время в сек
                "total_route_time_formatted": None # <--- НОВОЕ ПОЛЕ: общее время формат.
            }
        else:
            # Обновляем расчетные данные, сохраняя введенные пользователем
            self.summary[sanitized_name]["distance"] = distance_km if distance_km is not None else "Н/Д"
            self.summary[sanitized_name]["duration_seconds"] = duration_sec
            self.summary[sanitized_name]["duration_formatted"] = duration_formatted
            self.summary[sanitized_name]["number_of_stops"] = num_intermediate_stops # Обновляем кол-во точек

        # Пересчитываем все поля, включая новое "Время на маршруте"
        self._recalculate_summary_fields(sanitized_name)

    def get_route_names(self):
        return [data.get("original_name", key) for key, data in self.summary.items()]

    def get_route(self, name):
        sanitized_name = sanitize_filename(name)
        return self.routes.get(sanitized_name, {})

    def update_summary_item(self, name, report_distance=None, report_hours=None, report_minutes=None):
        """Обновляет данные для одного маршрута в сводке (без времени на точку)."""
        sanitized_name = sanitize_filename(name)
        if not sanitized_name or sanitized_name not in self.summary:
            print(f"⚠️ Попытка обновить несуществующий маршрут в сводке: {name} (sanitized: {sanitized_name})")
            return False

        summary_item = self.summary[sanitized_name]
        updated = False

        # Обновляем расстояние по отчету
        if report_distance is not None:
            try:
                report_distance_float = float(report_distance) if str(report_distance).strip() != "" else None
                if summary_item.get("report_distance") != report_distance_float:
                    summary_item["report_distance"] = report_distance_float
                    updated = True
            except (ValueError, TypeError):
                if summary_item.get("report_distance") is not None:
                     summary_item["report_distance"] = None # Сбрасываем если некорректный ввод
                     updated = True

        # --- Обновляем время по отчету ---
        new_report_h = None
        new_report_m = None
        try:
            if report_hours is not None and str(report_hours).strip() != "":
                new_report_h = int(report_hours)
            if report_minutes is not None and str(report_minutes).strip() != "":
                new_report_m = int(report_minutes)

            if summary_item.get("report_duration_hours") != new_report_h or \
               summary_item.get("report_duration_minutes") != new_report_m:
                summary_item["report_duration_hours"] = new_report_h
                summary_item["report_duration_minutes"] = new_report_m
                updated = True
                print(f"   Updated report time in summary_item: h={new_report_h}, m={new_report_m}")

        except (ValueError, TypeError):
            if summary_item.get("report_duration_hours") is not None or \
               summary_item.get("report_duration_minutes") is not None:
                summary_item["report_duration_hours"] = None
                summary_item["report_duration_minutes"] = None
                updated = True
                print(f"   Reset report time in summary_item due to conversion error")

        # Пересчитываем поля (включая время на маршруте) и сохраняем, если что-то обновилось
        if updated:
            self._recalculate_summary_fields(sanitized_name) # Пересчет включает новое поле
            self.save_to_disk()
            return True

        return False # Ничего не изменилось

    def _recalculate_summary_fields(self, sanitized_name, summary_dict=None):
        """Пересчитывает разницу в расстоянии, времени и общее время на маршруте."""
        target_summary = summary_dict if summary_dict is not None else self.summary

        if sanitized_name not in target_summary:
            return

        item = target_summary[sanitized_name]
        original_name = item.get("original_name", sanitized_name)
        print(f"-- Recalculating fields for: {original_name}")

        # --- Разница расстояний ---
        report_dist = item.get("report_distance")
        route_dist = item.get("distance")
        if report_dist is not None and isinstance(route_dist, (int, float)):
            try:
                item["distance_difference"] = round(route_dist - float(report_dist), 2)
            except (ValueError, TypeError):
                item["distance_difference"] = None
        else:
            item["distance_difference"] = None
        # print(f"   Distance diff calculated: {item['distance_difference']}") # Убрал лог для краткости

        # --- Расчет Общего Времени на Маршруте (ПРАВИЛЬНОЕ МЕСТО) ---
        total_route_time_seconds = None # Инициализируем переменную
        route_sec = item.get("duration_seconds") # Время в пути
        service_time_sec = 0
        if self.global_service_time_minutes is not None:
            try:
                # Используем .get с дефолтом 0 для number_of_stops
                service_time_sec = int(self.global_service_time_minutes) * 60 * item.get("number_of_stops", 0)
            except (ValueError, TypeError):
                 service_time_sec = 0

        if route_sec is not None:
            try:
                total_route_time_seconds = int(route_sec) + service_time_sec # <-- Расчет общего времени
                item["total_route_time_seconds"] = total_route_time_seconds # Сохраняем в item
                hours = total_route_time_seconds // 3600
                minutes = (total_route_time_seconds % 3600) // 60
                item["total_route_time_formatted"] = f"{hours} ч {minutes} мин" # Сохраняем формат
            except (ValueError, TypeError) as e:
                print(f"   Error calculating total route time for {original_name}: {e}")
                item["total_route_time_seconds"] = None
                item["total_route_time_formatted"] = None
        else:
            item["total_route_time_seconds"] = None
            item["total_route_time_formatted"] = None
            # print(f"   Skipping total route time calculation (route_sec={route_sec})") # Можно вернуть для отладки

        # --- Разница времени (отчет vs ОБЩЕЕ ВРЕМЯ НА МАРШРУТЕ) ---
        report_h = item.get("report_duration_hours")
        report_m = item.get("report_duration_minutes")
        # route_sec БОЛЬШЕ НЕ НУЖЕН ЗДЕСЬ, используем total_route_time_seconds

        # ... (логика расчета time_difference_formatted остается прежней) ...
        try:
            report_h_int = int(report_h) if report_h is not None and str(report_h).strip() != "" else 0
            report_m_int = int(report_m) if report_m is not None and str(report_m).strip() != "" else 0
            is_report_time_valid = not (report_h_int == 0 and report_m_int == 0 and (report_h is None or str(report_h).strip() == "") and (report_m is None or str(report_m).strip() == ""))
        except (ValueError, TypeError):
             is_report_time_valid = False

        calculate_time_diff = False
        # Проверяем, что ОБЩЕЕ время рассчитано
        if total_route_time_seconds is not None:
            if is_report_time_valid:
                 calculate_time_diff = True
            elif report_h_int == 0 and report_m_int == 0: # Если время отчета не введено (0ч 0м), тоже считаем разницу
                 calculate_time_diff = True # <-- ИЗМЕНЕНО: Считаем разницу, даже если время отчета 0, чтобы показать полное время маршрута как разницу

        if calculate_time_diff:
            try:
                report_total_seconds = report_h_int * 3600 + report_m_int * 60
                # ИСПОЛЬЗУЕМ ОБЩЕЕ ВРЕМЯ НА МАРШРУТЕ для расчета разницы
                time_diff_value_seconds = total_route_time_seconds - report_total_seconds # <-- РЕАЛЬНОЕ ЗНАЧЕНИЕ РАЗНИЦЫ
                sign = "+" if time_diff_value_seconds >= 0 else "-"
                diff_seconds_abs = abs(time_diff_value_seconds) # Используем abs() для форматирования
                diff_hours = diff_seconds_abs // 3600
                diff_minutes = (diff_seconds_abs % 3600) // 60
                item["time_difference_formatted"] = f"{sign}{diff_hours} ч {diff_minutes} мин"
                item["time_difference_seconds"] = time_diff_value_seconds # <-- СОХРАНЯЕМ ЗНАЧЕНИЕ В СЕКУНДАХ
            except (ValueError, TypeError) as e:
                print(f"   Error calculating time diff for {original_name}: {e}")
                item["time_difference_formatted"] = None
                item["time_difference_seconds"] = None # <-- Сбрасываем и секунды
        else:
            item["time_difference_formatted"] = None
            item["time_difference_seconds"] = None # <-- Сбрасываем и секунды

    def get_summary(self):
        try:
            with open(os.path.join(DATA_FOLDER, "original_route_names.json"), 'r', encoding='utf-8') as f:
                original_names_data = json.load(f)
                original_names = original_names_data.get("routes", [])
        except Exception as e:
            print(f"⚠️ Не удалось загрузить original_route_names.json: {e}")
            original_names = []

        # Сохраняем введенные пользовательские значения из текущего self.summary
        saved_user_input = {}
        for norm_name, data in self.summary.items():
            saved_user_input[norm_name] = {
                "report_distance": data.get("report_distance"),
                "report_duration_hours": data.get("report_duration_hours"),
                "report_duration_minutes": data.get("report_duration_minutes")
            }

        # Очищаем текущий summary перед заполнением актуальными данными
        current_summary = {} # Используем временный словарь для новых данных
        print("--- Обновление данных для сводки --- ")

        # Обрабатываем каждый маршрут из списка актуальных
        for original_name in original_names:
            sanitized_name = sanitize_filename(original_name)
            if not sanitized_name: continue
            
            print(f"  Обработка: {original_name}...")
            try:
                # Запрашиваем свежие данные маршрута (расстояние, точки и т.д.)
                route_data_response = get_route_data_endpoint(original_name)
                
                if route_data_response and not route_data_response.get("error"):
                    distance_data = route_data_response.get("distance_data", {})
                    distance_km = distance_data.get("total_distance") # Ожидаем км
                    duration_sec = distance_data.get("total_duration") # Ожидаем сек
                    duration_formatted = distance_data.get("formatted_duration", "Н/Д")
                    number_of_stops = route_data_response.get("number_of_stops", 0)
                    
                    # --- ПОЛУЧЕНИЕ ФИО ВОДИТЕЛЯ --- 
                    driver_name = self.drivers.get(sanitized_name, "—") # Получаем из self.drivers, дефолт "—"
                    print(f"    -> Водитель для {original_name}: {driver_name}")
                    # -----------------------------------
                    
                    # Создаем запись в НОВОЙ сводке
                    current_summary[sanitized_name] = {
                        "original_name": original_name,
                        "driver_name": driver_name, # <--- ДОБАВЛЕНО ПОЛЕ
                        "distance": distance_km if distance_km is not None else "Н/Д",
                        "duration_seconds": duration_sec,
                        "duration_formatted": duration_formatted,
                        "number_of_stops": number_of_stops, # Используем актуальное кол-во
                        "report_distance": None, # Будут восстановлены ниже
                        "report_duration_hours": None,
                        "report_duration_minutes": None,
                        "distance_difference": None,
                        "time_difference_formatted": None,
                        "time_difference_seconds": None, # <-- ДОБАВЛЕНО
                        "total_route_time_seconds": None, # Будут рассчитаны ниже
                        "total_route_time_formatted": None
                    }

                    # Восстанавливаем сохраненные пользовательские значения отчета
                    if sanitized_name in saved_user_input:
                        current_summary[sanitized_name]["report_distance"] = saved_user_input[sanitized_name]["report_distance"]
                        current_summary[sanitized_name]["report_duration_hours"] = saved_user_input[sanitized_name]["report_duration_hours"]
                        current_summary[sanitized_name]["report_duration_minutes"] = saved_user_input[sanitized_name]["report_duration_minutes"]
                        
                    # --- ПРИНУДИТЕЛЬНЫЙ ПЕРЕСЧЕТ ПОЛЕЙ ДЛЯ КАЖДОЙ ЗАПИСИ --- 
                    self._recalculate_summary_fields(sanitized_name, summary_dict=current_summary) # Передаем словарь для расчета
                        
                    print(f"    ✅ Данные для сводки получены и пересчитаны: {original_name}")
                else:
                    error_msg = route_data_response.get('message') if route_data_response else 'Данные не получены'
                    print(f"    ⚠️ Ошибка получения данных для маршрута {original_name}: {error_msg}")
            except Exception as e:
                print(f"    ❌ Исключение при обработке маршрута {original_name} для сводки: {e}")
                # import traceback
                # traceback.print_exc()

        # Обновляем основной summary новыми данными и сохраняем
        self.summary = current_summary
        self.save_to_disk() 
        print("--- Обновление сводки завершено --- ")
        return self.summary

    def save_to_disk(self):
         data = {
             "routes": self.routes,
             "summary": self.summary,
             "drivers": self.drivers, # <--- СОХРАНЯЕМ
             "current_file": self.current_file,
             "global_service_time_minutes": self.global_service_time_minutes # <--- СОХРАНЯЕМ
         }
         # Очищаем данные перед сохранением (убедимся, что функция работает корректно)
         # sanitized_data = sanitize_data_for_json(data) # Эта функция не нужна, json.dump сам справится с None
         with open(os.path.join(DATA_FOLDER, "route_data.json"), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2) # Используем data напрямую

    def load_from_disk(self):
        try:
            with open(os.path.join(DATA_FOLDER, "route_data.json"), "r", encoding="utf-8") as f:
                data = json.load(f)
                self.routes = data.get("routes", {})
                self.summary = data.get("summary", {})
                self.drivers = data.get("drivers", {}) # <--- ЗАГРУЖАЕМ
                self.current_file = data.get("current_file", "")
                # Загружаем глобальное время, или 0 если нет
                self.global_service_time_minutes = data.get("global_service_time_minutes", 0) # <--- ЗАГРУЖАЕМ

                print(f"💾 Данные загружены. Маршрутов: {len(self.routes)}, Записей в сводке: {len(self.summary)}, Водителей: {len(self.drivers)}, Время на точку: {self.global_service_time_minutes} мин.")

                # Миграция старых данных (добавляем новые поля в summary, если их нет)
                needs_recalculation = False
                for key, summary_item in self.summary.items():
                    if summary_item.setdefault("number_of_stops", None) is None:
                         # Пытаемся восстановить кол-во точек из routes, если возможно
                         route_entry = self.routes.get(key)
                         if route_entry and "route_points" in route_entry:
                             summary_item["number_of_stops"] = len(route_entry["route_points"])
                             needs_recalculation = True
                             print(f"   Восстановлено кол-во точек для {key}: {summary_item['number_of_stops']}")
                         else:
                             summary_item["number_of_stops"] = 0 # Или ставим 0, если не нашли
                             print(f"   Не удалось восстановить кол-во точек для {key}, установлено 0")
                    # Добавляем остальные поля со значением None, если их нет
                    if summary_item.setdefault("total_route_time_seconds", None) is None: needs_recalculation = True
                    if summary_item.setdefault("total_route_time_formatted", None) is None: needs_recalculation = True
                    # Убираем старое поле, если оно есть (на всякий случай)
                    summary_item.pop("service_time_per_stop_minutes", None)

                    # Убедимся, что обязательные поля для расчета существуют
                    summary_item.setdefault("report_duration_hours", None)
                    summary_item.setdefault("report_duration_minutes", None)
                    summary_item.setdefault("time_difference_formatted", None)
                    summary_item.setdefault("duration_seconds", None)
                    summary_item.setdefault("distance_difference", None)


                # Если добавили новые поля или восстановили кол-во точек, пересчитаем все
                if needs_recalculation:
                    print("   Некоторые поля отсутствовали или были восстановлены, запускаем полный пересчет сводки...")
                    for key in list(self.summary.keys()): # Используем list, чтобы не менять словарь во время итерации
                        self._recalculate_summary_fields(key)
                    self.save_to_disk() # Сохраняем пересчитанные данные

                return True
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"⚠️ Не удалось загрузить route_data.json: {e}. Начинаем с пустых данных.")
            self.routes = {}
            self.summary = {}
            self.current_file = ""
            self.global_service_time_minutes = 0 # Сброс глобального времени
            return False
        except Exception as e: # Ловим другие возможные ошибки при загрузке/миграции
             print(f"❌ Непредвиденная ошибка при загрузке/миграции данных route_data.json: {e}")
             # import traceback # Раскомментировать для детальной отладки
             # traceback.print_exc()
             self.routes = {}
             self.summary = {}
             self.current_file = ""
             self.global_service_time_minutes = 0
             return False

    # <--- НОВЫЙ МЕТОД --->
    def set_global_service_time(self, minutes):
        """Устанавливает глобальное время на точку и пересчитывает все маршруты."""
        try:
            # Преобразуем в int, если не None и не пустая строка, иначе 0
            new_time = int(minutes) if minutes is not None and str(minutes).strip() != "" else 0
            if new_time < 0: new_time = 0 # Время не может быть отрицательным

            if new_time != self.global_service_time_minutes:
                print(f"🔄 Установка нового глобального времени на точку: {new_time} мин.")
                self.global_service_time_minutes = new_time
                # Пересчитываем все маршруты в сводке
                recalculated_count = 0
                for key in list(self.summary.keys()):
                    self._recalculate_summary_fields(key)
                    recalculated_count += 1
                print(f"   Пересчитано маршрутов: {recalculated_count}")
                self.save_to_disk() # Сохраняем изменения
                return True
            else:
                print(f"ℹ️ Глобальное время на точку не изменилось ({new_time} мин).")
                return False
        except (ValueError, TypeError):
            print(f"⚠️ Некорректное значение для времени на точку: {minutes}. Оставляем текущее: {self.global_service_time_minutes} мин.")
            return False

# Создаем глобальный объект для хранения данных
route_data = RouteData()
# Пробуем загрузить данные из файла
if not route_data.load_from_disk():
    print("💾 Инициализирован пустой объект RouteData.")

@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    # Принудительно обрабатываем маршрут "Липецк 2" при первом запуске
    # Закомментировано для предотвращения автоматической обработки
    # if not os.path.exists(os.path.join(os.path.dirname(BASE_DIR), "geocoded_results_Липецк_2.csv")):
    #     print("⚠️ Принудительно обрабатываем маршрут 'Липецк 2'...")
    #     process_route("Липецк 2")
    
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/api/health")
def health_check():
    return {"status": "ok"}

@app.get("/api/routes")
def get_routes():
    try:
        with open(os.path.join(DATA_FOLDER, "original_route_names.json"), 'r', encoding='utf-8') as f:
            original_names_data = json.load(f)
            return {"routes": sorted(original_names_data.get("routes", []))}
    except Exception as e:
        print(f"⚠️ Не удалось загрузить original_route_names.json: {e}")
        return {"routes": []}

# Функция для обработки отдельного маршрута
def process_route(original_route_name):
    """Обработка маршрута с сохранением оригинального имени"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(script_dir)
    
    # Используем глобальные константы для путей
    file_name = sanitize_filename(original_route_name)
    parsed_path = os.path.join(PROJECT_DATA_FOLDER, "parsed_addresses", f"parsed_addresses_{file_name}.csv")
    geocoded_path = os.path.join(GEOCODED_RESULTS_FOLDER, f"geocoded_results_{file_name}.csv")
    
    if not os.path.exists(parsed_path):
        print(f"⚠️ Файл {parsed_path} не найден для маршрута {original_route_name}")
        return False
    
    process_success = False # Флаг успешности ВСЕГО процесса
    try:
        # Геокодирование
        with open(parsed_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        geocoded_output = []
        for row in rows:
            address = row["normalized_address"]
            result = geocode_address(address)
            result["excel_row"] = row["excel_row"]
            result["route_name"] = original_route_name
            geocoded_output.append(result)
        
        # Создаем директорию, если не существует
        os.makedirs(os.path.dirname(geocoded_path), exist_ok=True)
        
        with open(geocoded_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["excel_row", "route_name", "input", "found", "type", "description", "lat", "lon", "error"])
            writer.writeheader()
            writer.writerows(geocoded_output)
        print(f"✅ Геокодирование завершено для маршрута {original_route_name}")
        
        # Расчет расстояний
        print(f"\n🚀 Запускаем расчет расстояний для {original_route_name}...")
        # Вызываем импортированную функцию напрямую
        calculation_success = route_distance.calculate_and_save_route(
            route_name=original_route_name, 
            geocoded_file_path=geocoded_path
            # traffic_mode можно будет брать из config, если нужно
        )
        
        if calculation_success:
            print(f"✅ Расчет расстояний для {original_route_name} завершен успешно (через функцию).")
        else:
            print(f"❌ Ошибка при расчете расстояний для {original_route_name} (через функцию).")
            # Можно решить, прерывать ли весь процесс или нет
            # return False # Прервать, если расчет расстояний критичен

        # Добавление в route_data
        if calculation_success: # Добавляем только если расчет прошел
            try:
                route_info = get_route_data_endpoint(original_route_name)
                if route_info and not route_info.get("error"):
                    route_data.add_route(original_route_name, route_info)
                    route_data.save_to_disk()
                    print(f"✅ Маршрут {original_route_name} добавлен в сводку")
                    process_success = True # Весь процесс успешен
                else:
                    error_msg = route_info.get("message", "Не удалось получить данные маршрута") if route_info else "Не удалось получить данные маршрута"
                    print(f"⚠️ Не удалось добавить маршрут {original_route_name} в сводку: {error_msg}")
            except Exception as e:
                print(f"❌ Ошибка при добавлении маршрута {original_route_name} в сводку: {str(e)}")
        else:
             print(f"ℹ️ Маршрут {original_route_name} не добавлен в сводку, т.к. расчет расстояний не удался.")

    except Exception as e:
        print(f"❌ Общая ошибка при обработке маршрута {original_route_name}: {str(e)}")

    return process_success

@app.get("/api/route-data/{route_name}")
def get_route_data_endpoint(route_name: str = Path(...)):
    original_route_name = route_name
    if not original_route_name:
        raise HTTPException(status_code=400, detail="Имя маршрута не указано")

    try:
        logging.info(f"Getting route data for '{original_route_name}'")
        file_name = sanitize_filename(original_route_name)
        geocoded_file = os.path.join(GEOCODED_RESULTS_FOLDER, f"geocoded_results_{file_name}.csv")
        route_results_file = os.path.join(ROUTE_RESULTS_FOLDER, f"route_results_{file_name}.json")
        summary_item = route_data.summary.get(file_name, {}) # Получаем заранее

        # Ищем файл с результатами
        distance_data = {}
        if os.path.exists(route_results_file):
            logging.info(f"Found route results file: {route_results_file}")
            try:
                with open(route_results_file, 'r', encoding='utf-8') as f:
                    distance_data = json.load(f)

                # Форматируем данные для фронтенда
                if 'total_distance' in distance_data and distance_data['total_distance'] is not None:
                    try:
                        distance_meters = float(distance_data['total_distance'])
                        distance_km_rounded = round(distance_meters / 1000)
                        distance_data['formatted_distance'] = f"{distance_km_rounded} км"
                        distance_data['total_distance'] = distance_km_rounded
                    except (ValueError, TypeError):
                        print(f"⚠️ Не удалось преобразовать total_distance: {distance_data['total_distance']}")
                        distance_data['formatted_distance'] = "Ошибка"
                        distance_data['total_distance'] = None
                else:
                    distance_data['formatted_distance'] = "Н/Д"
                    distance_data['total_distance'] = None

                if 'total_duration' in distance_data and distance_data['total_duration'] is not None:
                    try:
                        total_seconds = int(distance_data['total_duration'])
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        formatted_duration = f"{hours} ч {minutes} мин" if hours > 0 else f"{minutes} мин"
                        distance_data['formatted_duration'] = formatted_duration
                        distance_data['total_duration'] = total_seconds
                    except (ValueError, TypeError):
                         print(f"⚠️ Не удалось преобразовать total_duration: {distance_data['total_duration']}")
                         distance_data['formatted_duration'] = "Ошибка"
                         distance_data['total_duration'] = None
                else:
                    distance_data['formatted_duration'] = "Н/Д"
                    distance_data['total_duration'] = None
            except Exception as e:
                logging.error(f"Error reading route results for {original_route_name}: {e}")
                distance_data = {"error": True, "error_message": f"Ошибка чтения файла результатов: {str(e)}"}
        else:
            logging.warning(f"Route results file not found for route {original_route_name}")
            distance_data = {"error": True, "error_message": f"Route results file not found for {original_route_name}"}


        # Чтение geocoded_data
        geocoded_data = []
        route_points = [] # Будет содержать только точки с валидными lat/lon
        number_of_stops_actual = 0
        try:
            if os.path.exists(geocoded_file):
                 geocoded_df = pd.read_csv(geocoded_file)
                 geocoded_df.replace([np.nan, np.inf, -np.inf], None, inplace=True)
                 for _, row in geocoded_df.iterrows():
                     point_data = row.to_dict()
                     geocoded_data.append(point_data) # Добавляем все строки для таблицы
                     # Добавляем в route_points только если есть валидные координаты
                     if 'lon' in point_data and 'lat' in point_data and point_data['lon'] is not None and point_data['lat'] is not None:
                        try:
                             lat = float(point_data['lat'])
                             lon = float(point_data['lon'])
                             if -90 <= lat <= 90 and -180 <= lon <= 180:
                                route_points.append({
                                     'lon': lon,
                                     'lat': lat,
                                     # Используем description для адреса на карте, input для исходного
                                     'address': point_data.get('description', point_data.get('input', '')),
                                     'original_address': point_data.get('input', '')
                                 })
                        except (ValueError, TypeError):
                             print(f"⚠️ Невалидные координаты в geocoded_file для {original_route_name}: {point_data.get('lat')}, {point_data.get('lon')}")


                 number_of_stops_actual = len(route_points) # Считаем только точки с координатами

                 # Обновляем количество точек в summary, если оно расходится и пересчитываем
                 if file_name in route_data.summary: 
                     if route_data.summary[file_name].get("number_of_stops") != number_of_stops_actual:
                         print(f"⚠️ Обновление кол-ва точек для {original_route_name} с {route_data.summary[file_name].get('number_of_stops')} на {number_of_stops_actual}")
                         route_data.summary[file_name]["number_of_stops"] = number_of_stops_actual
                         route_data._recalculate_summary_fields(file_name)
                         route_data.save_to_disk() # Сохраняем изменение
                         summary_item = route_data.summary.get(file_name, {}) # Обновляем локальный summary_item
                 else:
                     logging.warning(f"Route {original_route_name} (sanitized: {file_name}) not found in summary data.")

            else:
                 logging.warning(f"Could not find geocoded file for route {original_route_name}")
                 number_of_stops_actual = 0
                 geocoded_data = []
                 route_points = []
        except Exception as e:
             logging.error(f"Error reading geocoding data for {original_route_name}: {e}")
             geocoded_data = []
             route_points = []
             number_of_stops_actual = 0

        # --- ДОБАВЛЕНИЕ ОФИСА РТК В НАЧАЛО И КОНЕЦ СПИСКОВ ДЛЯ ФРОНТЕНДА ---
        try:
            office_loc = config.OFFICE_LOCATION
            if office_loc and "lat" in office_loc and "lon" in office_loc:
                office_lat = office_loc["lat"]
                office_lon = office_loc["lon"]
                office_name = office_loc.get("name", "Офис РТК")

                # Представление для таблицы geocoded_data
                office_geocoded_repr = {
                    "excel_row": "СТАРТ",
                    "route_name": original_route_name,
                    "input": office_name,
                    "found": True,
                    "type": "office",
                    "description": office_name,
                    "lat": office_lat,
                    "lon": office_lon,
                    "error": None
                }
                # Представление для карты route_points
                office_route_point_repr = {
                    "lat": office_lat,
                    "lon": office_lon,
                    "address": office_name,
                    "original_address": office_name
                }

                # Добавляем в начало
                geocoded_data.insert(0, office_geocoded_repr)
                route_points.insert(0, office_route_point_repr)

                # Добавляем в конец (копируем, чтобы excel_row можно было изменить)
                office_geocoded_repr_end = office_geocoded_repr.copy()
                office_geocoded_repr_end["excel_row"] = "ФИНИШ"
                geocoded_data.append(office_geocoded_repr_end)
                route_points.append(office_route_point_repr) # Координаты те же

                print(f"✅ Добавлен офис РТК в начало/конец данных для маршрута {original_route_name}")

            else:
                print("⚠️ Координаты OFFICE_LOCATION не найдены или неполны в config.py. Офис не добавлен.")
        except Exception as e_office:
            print(f"❌ Ошибка при добавлении офиса РТК: {e_office}")
        # --- КОНЕЦ ДОБАВЛЕНИЯ ОФИСА ---


        # Возвращаем данные, включая оригинальное имя маршрута и новые поля
        return {
            "route_name": original_route_name,
            "geocoder_output": geocoded_data, # <--- Теперь с офисом
            "route_points": route_points,     # <--- Теперь с офисом
            "distance_data": distance_data,
            "number_of_stops": summary_item.get("number_of_stops", number_of_stops_actual), 
            "total_route_time_formatted": summary_item.get("total_route_time_formatted", "Н/Д"),
            "global_service_time_minutes": route_data.global_service_time_minutes
        }
    except Exception as e:
        logging.error(f"Error in get_route_data_endpoint for {original_route_name}: {e}")
        # import traceback
        # traceback.print_exc()
        return {"error": True, "message": str(e), "route_name": original_route_name}

@app.get("/api/summary")
def get_summary_endpoint():
    summary = route_data.get_summary() # Этот метод теперь сам пересчитывает и сохраняет
    result = []
    # Перебираем актуальный route_data.summary после вызова get_summary()
    for norm_name, data in route_data.summary.items():
        result.append({
            "route_name": data.get("original_name", norm_name),
            "driver_name": data.get("driver_name", "—"),
            "distance": data.get("distance", "Н/Д"),
            "duration_seconds": data.get("duration_seconds"), # Время в пути (сек)
            "duration": data.get("duration_formatted", "Н/Д"), # Время в пути (формат)
            "report_distance": data.get("report_distance"),
            "report_duration_hours": data.get("report_duration_hours"),
            "report_duration_minutes": data.get("report_duration_minutes"),
            "distance_difference": data.get("distance_difference"),
            "time_difference_formatted": data.get("time_difference_formatted"),
            "time_difference_seconds": data.get("time_difference_seconds"), # <-- ДОБАВЛЕНО
            "total_route_time_seconds": data.get("total_route_time_seconds"), # <-- Добавляем общее время в секундах
            "total_route_time_formatted": data.get("total_route_time_formatted", "Н/Д"), # Время на маршруте (формат)
            "number_of_stops": data.get("number_of_stops", "Н/Д") # <--- ДОБАВЛЕНО КОЛ-ВО ТОЧЕК
        })
    # Добавляем глобальное время на точку в ответ
    return {
        "summary": result, # <<< УБРАНА СОРТИРОВКА
        "global_service_time_minutes": route_data.global_service_time_minutes
     }

@app.post("/api/upload")
async def upload_excel(file: UploadFile, service_time_per_stop_minutes: int = Form(0)): # <--- ИЗМЕНЕНА СИГНАТУРА
    save_path = os.path.join(UPLOAD_FOLDER, file.filename)
    with open(save_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # Сохраняем имя текущего файла
    route_data.current_file = file.filename

    # --- Печатаем полученное значение ---
    print(f"--- Received time per stop in /api/upload: {service_time_per_stop_minutes} (type: {type(service_time_per_stop_minutes)}) ---")
    # --- Конец блока ---

    # --- Устанавливаем глобальное время на точку ---
    # Убираем проверку на None, так как параметр теперь всегда int
    # time_to_set = service_time_per_stop_minutes if service_time_per_stop_minutes is not None else 0
    route_data.set_global_service_time(service_time_per_stop_minutes)
    # --- Конец блока ---

    try:
        # --- Чтение Excel и определение маршрутов --- 
        df = pd.read_excel(save_path)
        all_routes_addresses = {} # Словарь: route_name -> [(row_num, address), ...]
        route_names = []          # Список имен маршрутов в порядке их появления
        current_route = None
        seen_kontragents_in_route = set() # <-- Множество для отслеживания контрагентов в ТЕКУЩЕМ маршруте
        
        # --- Добавлена проверка на пустой файл --- 
        if df.empty:
             raise ValueError("Загруженный Excel файл пуст.")
        # --- Конец проверки --- 

        # --- Ищем колонки динамически (аналогично parsing_route.py) --- 
        region_col_name = None
        address_col_name = "Адрес доставки"
        kontragent_col_name = None # <-- Имя колонки контрагентов
        driver_col_name = None # <--- Имя колонки водителя
        
        for col in df.columns:
            col_str = str(col).strip()
            if region_col_name is None and col_str.startswith("Регион/Маршрут"):
                region_col_name = col_str
            if kontragent_col_name is None and col_str.startswith("Контрагентов"):
                 kontragent_col_name = col_str
            if driver_col_name is None and col_str.startswith("Водитель"): # <--- Ищем колонку Водитель
                 driver_col_name = col_str
                
        # Проверяем наличие всех необходимых колонок
        missing_cols = []
        if region_col_name is None: missing_cols.append("колонка, начинающаяся с 'Регион/Маршрут'")
        if address_col_name not in df.columns: missing_cols.append(f"колонка '{address_col_name}'")
        if kontragent_col_name is None: missing_cols.append("колонка, начинающаяся с 'Контрагентов'")
        if driver_col_name is None: missing_cols.append("колонка, начинающаяся с 'Водитель'") # <--- Проверяем колонку Водитель
        
        if missing_cols:
             raise ValueError(f"Отсутствуют необходимые колонки: { ' и '.join(missing_cols) }.")
        print(f"  ✅ Найдены колонки: '{region_col_name}', '{address_col_name}', '{kontragent_col_name}', '{driver_col_name}'") # <--- Добавили в лог
        # --- Конец динамического поиска --- 
        
        # Сбрасываем старые данные о водителях перед обработкой нового файла
        route_data.drivers = {}
        print("  ℹ️ Словарь водителей очищен перед обработкой нового файла.")

        for idx, row in df.iterrows():
            # Используем найденные имена колонок
            region = row.get(region_col_name)
            address = row.get(address_col_name)
            kontragent = row.get(kontragent_col_name) # <-- Получаем контрагента
            driver = row.get(driver_col_name) # <--- Получаем водителя
            
            # Определяем начало нового маршрута
            is_new_route_marker = False
            if pd.notna(region) and isinstance(region, str) and region.strip():
                potential_new_route = region.strip()
                if pd.isna(address) or not str(address).strip() or current_route is None or potential_new_route != current_route:
                     is_new_route_marker = True
                     current_route = potential_new_route
                     if current_route not in route_names: # Сохраняем порядок только для НЕПУСТЫХ маршрутов
                         route_names.append(current_route)
                     if current_route not in all_routes_addresses:
                         all_routes_addresses[current_route] = []
                     seen_kontragents_in_route.clear() # <-- Очищаем сет контрагентов для нового маршрута
                     # driver_found_for_current_route = False # <-- Сбрасываем флаг для нового маршрута (Более простой способ - проверка наличия ключа)
            
            # Добавляем адрес к ТЕКУЩЕМУ маршруту, если:
            # 1. Маршрут определен
            # 2. Адрес есть
            # 3. Контрагент есть и он еще НЕ встречался в ЭТОМ маршруте
            if current_route and pd.notna(address) and isinstance(address, str) and address.strip():
                excel_row = idx + 2 # +1 за 0-индекс, +1 за строку заголовка
                if pd.notna(kontragent) and isinstance(kontragent, str) and kontragent.strip():
                    kontragent_key = kontragent.strip()
                    if kontragent_key not in seen_kontragents_in_route:
                        seen_kontragents_in_route.add(kontragent_key) # Добавляем контрагента в сет
                        all_routes_addresses[current_route].append((excel_row, address.strip()))
                        
                        # --- СОХРАНЕНИЕ ВОДИТЕЛЯ --- 
                        sanitized_route_name = sanitize_filename(current_route)
                        # Проверяем, что водитель еще не найден для этого маршрута И что значение водителя валидно
                        if sanitized_route_name and sanitized_route_name not in route_data.drivers:
                            if pd.notna(driver) and isinstance(driver, str) and driver.strip():
                                driver_name = driver.strip()
                                route_data.drivers[sanitized_route_name] = driver_name
                                print(f"    🧑 Найден водитель для '{current_route}': {driver_name}")
                        # --- КОНЕЦ СОХРАНЕНИЯ ВОДИТЕЛЯ ---
                        
                    # else: # Логгирование дубликата (опционально)
                        # print(f"  [upload] Пропуск строки {excel_row}: дубль контрагента '{kontragent_key}' в '{current_route}'")
                # else: # Логгирование отсутствующего контрагента (опционально)
                     # print(f"  [upload] Пропуск строки {excel_row}: нет контрагента в '{current_route}'")
        
        # Удаляем маршруты без адресов из обоих списков
        route_names = [r for r in route_names if all_routes_addresses.get(r)]
        all_routes_addresses = {r: adds for r, adds in all_routes_addresses.items() if adds}

        if not route_names:
            raise ValueError("Не найдено ни одного маршрута с адресами в файле.")
            
        print(f"✅ В файле найдено маршрутов с адресами: {len(route_names)}")
        
        # --- Сохраняем route_data с обновленными водителями ПЕРЕД запуском парсинга --- 
        if route_data.drivers: # Сохраняем, только если нашли хотя бы одного водителя
            print("💾 Сохраняем route_data с информацией о водителях...")
            route_data.save_to_disk()
        # -----------------------------------------------------------------------------
        
        # --- Очистка старых файлов и кеша для маршрутов из нового файла --- 
        print("\n🧹 Начинаем очистку старых данных для маршрутов из файла...")
        routes_cleaned_count = 0
        # Добавляем очистку summary CSV, если он существует
        summary_csv_path = os.path.join(SUMMARY_CSV_FOLDER, "current_summary.csv")
        # Очищаем его только если загрузка прошла успешно до этого момента
        # (Лучше очищать после успешной обработки всех маршрутов)
        # if os.path.exists(summary_csv_path):
        #     try:
        #         os.remove(summary_csv_path)
        #         print(f"  🗑️ Удален старый файл сводки: {os.path.basename(summary_csv_path)}")
        #     except OSError as e:
        #         print(f"  ❌ Ошибка при удалении файла {summary_csv_path}: {e}")

        for route_name_to_clean in route_names:
            file_name = sanitize_filename(route_name_to_clean)
            if not file_name:
                 continue
                 
            # Пути к файлам для удаления
            parsed_path = os.path.join(PARSED_ADDRESSES_FOLDER, f"parsed_addresses_{file_name}.csv")
            geocoded_path = os.path.join(GEOCODED_RESULTS_FOLDER, f"geocoded_results_{file_name}.csv")
            route_results_path = os.path.join(ROUTE_RESULTS_FOLDER, f"route_results_{file_name}.json")
            
            files_to_delete = [parsed_path, geocoded_path, route_results_path]
            deleted_something_for_route = False

            # Удаляем файлы
            for file_path in files_to_delete:
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        print(f"  🗑️ Удален файл: {os.path.basename(file_path)} (для '{route_name_to_clean}')")
                        deleted_something_for_route = True
                except OSError as e:
                    print(f"  ❌ Ошибка при удалении файла {file_path}: {e}")
            
            # Удаляем из кеша (summary и routes)
            cache_key = sanitize_filename(route_name_to_clean) # Используем то же самое имя файла
            route_popped = route_data.routes.pop(cache_key, None)
            summary_popped = route_data.summary.pop(cache_key, None)
            
            if route_popped is not None:
                 print(f"  🗑️ Удалена запись из route_data.routes для '{route_name_to_clean}'")
                 deleted_something_for_route = True
            if summary_popped is not None:
                 print(f"  🗑️ Удалена запись из route_data.summary для '{route_name_to_clean}'")
                 deleted_something_for_route = True
                 
            if deleted_something_for_route:
                routes_cleaned_count += 1

        if routes_cleaned_count > 0:
            print(f"💾 Сохраняем очищенный route_data...")
            route_data.save_to_disk() # Сохраняем изменения в кеше (удаление)
        else:
            print("ℹ️ Не найдено старых данных для очистки.")
            
        # Сохраняем ОРИГИНАЛЬНЫЕ имена маршрутов из ТЕКУЩЕГО файла
        routes_file = os.path.join(DATA_FOLDER, "original_route_names.json")
        with open(routes_file, 'w', encoding='utf-8') as f:
            json.dump({"routes": route_names}, f, ensure_ascii=False, indent=2)
        print(f"💾 Актуальный список маршрутов сохранен в {os.path.basename(routes_file)}")
        
        # --- Запуск парсинга и обработки для каждого маршрута --- 
        all_exceptions = []
        processed_route_names = [] # Сохраняем имена успешно обработанных маршрутов
        parsing_failed_routes = [] # Сохраняем имена маршрутов, где парсинг упал
        print("\n🚚 Запускаем парсинг и обработку маршрутов...")
        
        # --- ШАГ 1: Только парсинг и сбор исключений --- 
        for route_name in route_names:
            addresses_with_rows = all_routes_addresses[route_name]
            print(f"\n--- Парсинг маршрута: {route_name} ---")
            
            # 1. Запускаем парсинг через LLM
            print(f"  💬 Запуск parsing_route.py...")
            openrouter_key = os.getenv("OPENROUTER_API_KEY", "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")
            if not openrouter_key:
                print("  ❌ ОШИБКА: Ключ OpenRouter не найден (OPENROUTER_API_KEY).")
                raise ValueError("Ключ OpenRouter API не настроен.")

            parsing_cmd = [
                 sys.executable, 
                 os.path.abspath(os.path.join(BASE_DIR, "..", "parsing_route.py")),
                 "--excel", save_path,
                 "--openrouter_key", openrouter_key, 
                 "--model", "google/gemini-flash-1.5",
                 "--route", route_name
            ]
            cmd_str = ' '.join(parsing_cmd)
            print(f"  Выполняем команду: {cmd_str}")
            parsing_result = subprocess.run(parsing_cmd, capture_output=True, text=True, cwd=ROOT_DIR)
            
            if parsing_result.returncode != 0:
                print(f"  ❌ Ошибка выполнения parsing_route.py для '{route_name}'")
                print(f"  Stderr:\n{parsing_result.stderr}")
                parsing_failed_routes.append(route_name) # Запоминаем, что парсинг упал
                continue # Пропускаем сбор исключений для этого маршрута
            else:
                print(f"  ✅ parsing_route.py выполнен успешно.")
                # --- ОБНОВЛЕННАЯ ЛОГИКА ПАРСИНГА ИСКЛЮЧЕНИЙ ИЗ STDOUT --- 
                lines = parsing_result.stdout.splitlines()
                route_exceptions = [] # Используем этот список для сбора исключений
                
                for line in lines:
                    # Ищем строки, соответствующие НОВОМУ формату вывода исключений
                    match = re.match(r"Маршрут\s+'(.*?)',\s+Строка\s+(\d+):\s*(.*)", line.strip(), re.IGNORECASE)
                    if match:
                        route_nm_parsed = match.group(1).strip()
                        row_num_str = match.group(2)
                        original_address_parsed = match.group(3).strip()
                        
                        # Убираем возможные кавычки по краям адреса
                        if (original_address_parsed.startswith("'") and original_address_parsed.endswith("'")) or \
                           (original_address_parsed.startswith('"') and original_address_parsed.endswith('"')):
                            original_address_parsed = original_address_parsed[1:-1]
                        
                        try:
                            row_num = int(row_num_str)
                            # Ищем оригинальный адрес по номеру строки для ЭТОГО маршрута
                            # Важно: route_name здесь - это имя маршрута из текущей итерации внешнего цикла!
                            # Нужно найти правильный addresses_with_rows для route_nm_parsed
                            current_route_addresses_rows = all_routes_addresses.get(route_nm_parsed)
                            if current_route_addresses_rows:
                                original_address_excel = next((addr for r, addr in current_route_addresses_rows if r == row_num), original_address_parsed)
                            else:
                                original_address_excel = original_address_parsed # Если не нашли список адресов для маршрута
                                
                            route_exceptions.append({
                                "row": row_num,
                                "address": original_address_excel,
                                "route": route_nm_parsed, # Используем имя из строки лога
                                "corrected": ""
                            })
                            print(f"    ✅ [Из stdout] Распознано исключение: Маршрут '{route_nm_parsed}', Строка {row_num}, Адрес: '{original_address_excel}'")
                        except ValueError:
                            print(f"    ⚠️ [Из stdout] Не удалось распознать номер строки в исключении: {line.strip()}")
                        except Exception as e_inner:
                             print(f"    ⚠️ [Из stdout] Ошибка при обработке строки исключения '{line.strip()}': {e_inner}")
                             
                # Добавляем ВСЕ найденные исключения (из route_exceptions) в общий список
                all_exceptions.extend(route_exceptions)
                # --- КОНЕЦ ОБНОВЛЕННОЙ ЛОГИКИ ПАРСИНГА --- 

        # --- ШАГ 2: Проверка наличия исключений и решение --- 
        if all_exceptions:
            print("\n❗️ Найдены адреса, требующие проверки. Отправка исключений на фронтенд.")
            # Если есть исключения, НЕ запускаем геокодирование/расчет, возвращаем исключения
            return JSONResponse({
                "status": "needs_correction", 
                "exceptions": all_exceptions,
                "routes": route_names # Отправляем все имена маршрутов, чтобы селектор обновился
            })
        else:
            print("\n✅ Адреса не требуют ручной проверки. Запускаем геокодирование и расчет...")
            # Если исключений НЕТ, запускаем геокодирование и расчет для ВСЕХ маршрутов
            # (Пропускаем те, где упал парсинг, если такие были)
            for route_name in route_names:
                if route_name in parsing_failed_routes:
                    print(f"  ⚠️ Пропуск геокодирования/расчета для '{route_name}', т.к. парсинг не удался.")
                    continue
                
                print(f"\n--- Геокод+Расчет для: {route_name} ---")
                process_success = process_route(route_name) 
                if process_success:
                    print(f"  ✅ Маршрут '{route_name}' успешно обработан.")
                    processed_route_names.append(route_name)
                else:
                    print(f"  ⚠️ Маршрут '{route_name}' не был полностью обработан (ошибка в process_route).")
            
            # Очищаем CSV сводки только если были успешно обработанные маршруты
            if processed_route_names and os.path.exists(summary_csv_path):
                 try:
                     os.remove(summary_csv_path)
                     print(f"  🗑️ Удален старый файл сводки: {os.path.basename(summary_csv_path)} (т.к. есть новые данные)")
                 except OSError as e:
                     print(f"  ❌ Ошибка при удалении файла {summary_csv_path}: {e}")
                     
            print("\n🏁 Обработка всех маршрутов из файла завершена (без исключений).")
            # Возвращаем статус processed и список УСПЕШНО обработанных маршрутов
            return JSONResponse({
                "status": "processed", 
                "exceptions": [], # Пустой список исключений
                "routes": processed_route_names 
            })

    except ValueError as ve:
        print(f"❌ Ошибка при обработке файла (ValueError): {ve}")
        # import traceback # Для отладки
        # traceback.print_exc()
        return JSONResponse(status_code=400, content={"error": "Ошибка в данных Excel", "details": str(ve)})
    except subprocess.CalledProcessError as cpe:
        print(f"❌ Ошибка при вызове подпроцесса: {cpe}")
        print(f"   Stderr: {cpe.stderr}")
        return JSONResponse(status_code=500, content={"error": "Ошибка при обработке маршрута", "details": "Не удалось выполнить один из этапов обработки."}) 
    except Exception as e:
        print(f"❌ Непредвиденная ошибка при обработке файла: {e}")
        import traceback # Важно для отладки неожиданных ошибок
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": "Внутренняя ошибка сервера", "details": "Произошла ошибка при обработке файла."}) 

@app.post("/api/submit-corrections")
async def submit_corrections(data: dict = Body(...)):
    corrections = data.get("corrections", [])
    print("📥 Получены исправления:", corrections)

    # Создаем словарь для быстрого доступа к исправлениям по *нормализованному* route_name и excel_row
    corrections_map = {}
    original_names_map = {} # Для хранения оригинального имени по нормализованному
    routes_to_process_normalized = set() # Используем нормализованные имена для обработки

    for corr in corrections:
        original_route = corr.get("route", "")
        if not original_route:
            continue # Пропускаем исправления без имени маршрута
            
        normalized_route = sanitize_filename(original_route)
        if not normalized_route:
             continue # Пропускаем, если имя стало пустым после нормализации
             
        routes_to_process_normalized.add(normalized_route)
        original_names_map[normalized_route] = original_route # Запоминаем оригинальное имя
        
        row_num_str = str(corr["row"])
        if normalized_route not in corrections_map:
            corrections_map[normalized_route] = {}
        corrections_map[normalized_route][row_num_str] = corr["corrected"]

    # --- Loop 1: Modify parsed_addresses_*.csv files --- 
    processed_routes_in_loop1 = set()
    for normalized_route in routes_to_process_normalized:
        original_name = original_names_map[normalized_route]
        parsed_path = os.path.join(PARSED_ADDRESSES_FOLDER, f"parsed_addresses_{normalized_route}.csv")
        
        if not os.path.exists(parsed_path):
            print(f"⚠️ Файл {parsed_path} не найден для маршрута '{original_name}' в цикле 1")
            continue

        try:
            with open(parsed_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                fieldnames = reader.fieldnames or ["excel_row", "route_name", "normalized_address"]

            # Вносим правки для текущего нормализованного маршрута
            route_corrections_for_file = corrections_map.get(normalized_route, {})
            modified = False
            for row in rows:
                # Нормализуем имя маршрута из строки CSV для сравнения
                row_original_name = row.get('route_name')
                if sanitize_filename(row_original_name) == normalized_route:
                    row_num_str = row.get("excel_row")
                    if row_num_str in route_corrections_for_file:
                        corrected_address = route_corrections_for_file[row_num_str]
                        if row.get('normalized_address') != corrected_address:
                             print(f"✏️ [Файл {original_name}] Обновляем строку {row_num_str}: '{row.get('normalized_address')}' -> '{corrected_address}'")
                             row['normalized_address'] = corrected_address
                             modified = True
            
            # Сохраняем, только если были изменения
            if modified:
                with open(parsed_path, "w", encoding="utf-8", newline="") as f:
                      writer = csv.DictWriter(f, fieldnames=fieldnames)
                      writer.writeheader()
                      writer.writerows(rows)
                      print(f"💾 Файл {parsed_path} обновлен исправлениями.")
                 
            processed_routes_in_loop1.add(normalized_route) # Отмечаем нормализованное имя
            
        except Exception as e:
            print(f"❌ Ошибка при обновлении файла {parsed_path}: {e}")

    # --- Loop 2: Geocode and Calculate Distance FOR ALL ROUTES --- 
    print("\n--- Запуск геокодирования и расчета расстояний для ВСЕХ маршрутов после исправлений ---")
    
    # Загружаем ПОЛНЫЙ список оригинальных имен из файла
    try:
        with open(os.path.join(DATA_FOLDER, "original_route_names.json"), 'r', encoding='utf-8') as f:
            original_names_data = json.load(f)
            all_original_route_names = original_names_data.get("routes", [])
            print(f"  Будут обработаны маршруты: {', '.join(all_original_route_names)}")
    except Exception as e:
        print(f"❌ Не удалось загрузить полный список маршрутов из original_route_names.json: {e}")
        # Если не удалось загрузить, пытаемся обработать хотя бы исправленные
        all_original_route_names = [original_names_map[norm_name] for norm_name in routes_to_process_normalized if norm_name in original_names_map]
        print(f"  ⚠️ Используется список только исправленных маршрутов: {', '.join(all_original_route_names)}")
        if not all_original_route_names:
             return JSONResponse(status_code=500, content={"error": "Ошибка обработки", "details": "Не удалось определить список маршрутов для обработки после исправлений."}) 

    # Перебираем ВСЕ оригинальные имена из загруженного файла
    for original_name in all_original_route_names:
        normalized_route = sanitize_filename(original_name)
        if not normalized_route: continue # Пропускаем если имя некорректно

        parsed_path = os.path.join(PARSED_ADDRESSES_FOLDER, f"parsed_addresses_{normalized_route}.csv")
        geocoded_path = os.path.join(GEOCODED_RESULTS_FOLDER, f"geocoded_results_{normalized_route}.csv")

        if not os.path.exists(parsed_path):
            print(f"⚠️ Файл {os.path.basename(parsed_path)} не найден для маршрута '{original_name}' в цикле 2. Пропуск.")
            continue
        
        # Вызываем process_route, который делает и геокод, и расчет, и обновление route_data
        print(f"\n--- Обработка (process_route): {original_name} ---")
        try:
            process_success = process_route(original_name)
            if process_success:
                 print(f"✅ Маршрут '{original_name}' успешно обработан через process_route.")
            else:
                 print(f"⚠️ Маршрут '{original_name}' не был полностью обработан через process_route.")
        except Exception as e_proc:
            print(f"❌ Исключение при вызове process_route для '{original_name}': {e_proc}")
            # import traceback # Раскомментировать для детальной отладки
            # traceback.print_exc()

    # Получаем обновленный список маршрутов для возврата клиенту (используем оригинальные имена)
    try:
        routes_response = get_routes() # get_routes уже возвращает оригинальные имена
        routes_list = routes_response.get("routes", [])
    except Exception as e:
        print(f"❌ Ошибка при получении списка маршрутов: {str(e)}")
        routes_list = []
        
    # Возвращаем статус об успешном сохранении
    return JSONResponse(status_code=200, content={"status": "saved", "routes": routes_list})

@app.post("/api/summary/update")
async def update_summary_endpoint(data: dict = Body(...)):
    original_route_name = data.get("route_name")
    report_distance = data.get("report_distance")
    # Получаем поля времени отчета
    report_hours = data.get("report_duration_hours")
    report_minutes = data.get("report_duration_minutes")

    # !!! Убеждаемся, что service_time_per_stop_minutes НЕ принимается и НЕ передается в update_summary_item !!!
    # Это поле теперь глобальное и устанавливается через /api/upload

    # <<< Логгирование полученных данных >>>
    print(f"--- Received /api/summary/update for: {original_route_name}")
    print(f"    report_distance: '{report_distance}'")
    print(f"    report_duration_hours: '{report_hours}'")
    print(f"    report_duration_minutes: '{report_minutes}'")
    # <<< Конец логгирования >>>

    if not original_route_name:
        raise HTTPException(status_code=400, detail="Необходимо указать имя маршрута")

    # Обновляем данные в route_data, вызывая функцию ТОЛЬКО с данными отчета
    updated = route_data.update_summary_item(
        name=original_route_name,
        report_distance=report_distance,
        report_hours=report_hours,
        report_minutes=report_minutes
        # service_time_per_stop_minutes НЕ передается
    )

    if updated:
        # CSV пока не трогаем по вашему указанию
        # csv_path = save_summary_to_csv(route_data.summary)
        # print(f"✅ Данные сохранены в CSV: {csv_path}")
        print(f"✅ Данные для '{original_route_name}' обновлены в памяти.")
    else:
        print(f"ℹ️ Данные для маршрута '{original_route_name}' не изменились.")

    # Возвращаем обновленные данные для всей сводки
    summary_response = get_summary_endpoint() # Вызываем обновленный эндпоинт
    # Добавляем статус "ok" к ответу от get_summary_endpoint
    return {"status": "ok", **summary_response}

# --- НОВЫЙ ENDPOINT ДЛЯ ЭКСПОРТА СВОДКИ --- 
@app.post("/api/export-summary")
async def export_summary_endpoint():
    print("-- Export summary requested --")
    global route_data # <<< ИСПРАВЛЕНО: route_data вместо route_processor
    if not route_data or not route_data.summary: # <<< ИСПРАВЛЕНО
        print("   No summary data found for export.")
        raise HTTPException(status_code=404, detail="Нет данных для экспорта сводки.")

    try:
        # Используем метод get_summary() экземпляра route_data
        summary_data = route_data.get_summary() # <<< ИЗМЕНЕНО имя переменной и ИСПРАВЛЕНО
        print(f"   Got {len(summary_data)} items for summary export.")

        # Проверяем, что get_summary() вернул данные (он возвращает словарь)
        if not summary_data: 
             print("   get_summary() returned empty data.")
             raise HTTPException(status_code=404, detail="Нет данных для экспорта сводки (метод вернул пусто).")

        # Преобразуем словарь значений в список для DataFrame
        summary_values_list = list(summary_data.values())
        if not summary_values_list:
            print("   Summary dictionary is empty after getting values.")
            raise HTTPException(status_code=404, detail="Нет данных для экспорта сводки (словарь пуст).")

        # 1. Преобразуем список словарей в DataFrame
        df = pd.DataFrame(summary_values_list)

        # --- СОРТИРОВКА DataFrame по исходному порядку маршрутов --- 
        original_route_order = []
        try:
            with open(os.path.join(DATA_FOLDER, "original_route_names.json"), 'r', encoding='utf-8') as f:
                original_names_data = json.load(f)
                original_route_order = original_names_data.get("routes", [])
            
            if original_route_order and 'original_name' in df.columns:
                # Преобразуем колонку original_name в категориальный тип с нужным порядком
                df['original_name'] = pd.Categorical(df['original_name'], categories=original_route_order, ordered=True)
                df = df.sort_values('original_name')
                # Сбрасываем индекс после сортировки
                df = df.reset_index(drop=True)
                print(f"   DataFrame отсортирован по исходному порядку: {original_route_order}")
            else:
                print("   ⚠️ Не удалось отсортировать DataFrame по исходному порядку (список пуст или нет колонки original_name).")
        except Exception as e:
            print(f"   ⚠️ Ошибка при чтении/сортировке по original_route_names.json: {e}")
        # --- КОНЕЦ СОРТИРОВКИ --- 

        # --- ФОРМАТИРОВАНИЕ ВРЕМЕНИ ОТЧЕТА --- 
        def format_report_time(row):
            hours = row.get('report_duration_hours')
            minutes = row.get('report_duration_minutes')
            
            # Обрабатываем NaN/None как 0, но только если хотя бы одно значение не None
            has_hours = hours is not None and pd.notna(hours)
            has_minutes = minutes is not None and pd.notna(minutes)
            
            if not has_hours and not has_minutes:
                return None # Возвращаем None, если оба значения отсутствуют
                
            h = int(hours) if has_hours else 0
            m = int(minutes) if has_minutes else 0
            
            # Валидация (на всякий случай, хотя на бэке уже есть)
            if h < 0: h = 0
            if m < 0: m = 0
            if m >= 60: m = 59 
            
            return f"{h} ч {m} мин"

        # Применяем функцию для создания новой колонки
        if 'report_duration_hours' in df.columns and 'report_duration_minutes' in df.columns:
            df['report_duration_formatted'] = df.apply(format_report_time, axis=1)
            print("   Column 'report_duration_formatted' created.")
        else:
             print("   Warning: Columns 'report_duration_hours' or 'report_duration_minutes' not found in df. Skipping report time formatting.")
             df['report_duration_formatted'] = None # Создаем пустую колонку, чтобы избежать ошибок ниже
        # --- КОНЕЦ ФОРМАТИРОВАНИЯ --- 

        # --- Выбираем и переименовываем колонки --- 
        columns_to_export = {
            "original_name": "Маршрут",
            "driver_name": "ФИО водителя",
            "distance": "Расстояние (км)",
            "total_route_time_formatted": "Время на маршруте",
            "report_distance": "Расстояние (отчет)",
            "report_duration_formatted": "Время (отчет)", # <-- Используем новую колонку
            # "report_duration_hours": "Время (отчет) Часы", # <-- Удалено
            # "report_duration_minutes": "Время (отчет) Минуты", # <-- Удалено
            "distance_difference": "Разница (км)",
            "time_difference_formatted": "Разница (время)",
            "number_of_stops": "Кол-во точек"
        }
        
        # Проверяем наличие колонок перед экспортом
        export_df = pd.DataFrame()
        ordered_columns = [] # Сохраняем порядок колонок
        for original_col, new_name in columns_to_export.items():
            ordered_columns.append(new_name) # Добавляем в том порядке, как они в словаре
            if original_col in df.columns:
                export_df[new_name] = df[original_col]
            else:
                print(f"   Warning: Column '{original_col}' not found in summary data for export.")
                export_df[new_name] = None # Или pd.NA или другое значение по умолчанию
        
        # Упорядочиваем колонки DataFrame
        export_df = export_df[ordered_columns]

        # --- Очистка ячеек разницы, если нет данных отчета --- 
        # Заменяем NaN и None на пустые строки для корректной проверки pd.isna()
        # Нужно сделать копию, чтобы избежать SettingWithCopyWarning
        export_df_copy = export_df.copy()

        # Для расстояния
        export_df_copy.loc[pd.isna(export_df_copy['Расстояние (отчет)']), 'Разница (км)'] = None

        # Для времени (проверяем исходные колонки в df)
        report_time_is_empty_in_source = pd.isna(df['report_duration_hours']) & pd.isna(df['report_duration_minutes'])
        # Применяем маску к export_df_copy, используя индексы из df
        export_df_copy.loc[report_time_is_empty_in_source, 'Разница (время)'] = None
        
        # Используем измененную копию для экспорта
        export_df = export_df_copy 
        # --- Конец очистки --- 

        # Можно добавить форматирование, если нужно
        # export_df['Разница (км)'] = export_df['Разница (км)'].apply(lambda x: f"{'+' if x>0 else ''}{x:.2f}" if pd.notna(x) else '')
        
        print(f"   DataFrame prepared for export with columns: {list(export_df.columns)}")

        # 2. Создаем Excel в памяти
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Сводка')
            
            # --- УСЛОВНОЕ ФОРМАТИРОВАНИЕ СТРОК --- 
            workbook = writer.book
            worksheet = writer.sheets['Сводка']
            
            # Определяем стиль заливки (светло-красный)
            light_red_fill = PatternFill(start_color='FFFFCCCC',
                                       end_color='FFFFCCCC',
                                       fill_type='solid')
                                       
            # Константы для порогов
            DISTANCE_THRESHOLD = -40
            TIME_THRESHOLD_SECONDS = -5400 # -1 час 30 минут
            
            # Итерация по строкам данных листа Excel (пропускаем заголовок - row 1)
            for row_idx in range(2, worksheet.max_row + 1):
                # Получаем индекс для исходного DataFrame df (который отсортирован)
                df_idx = row_idx - 2
                
                try:
                    # Получаем необработанные значения разницы из исходного df
                    dist_diff_val = df.loc[df_idx, 'distance_difference']
                    time_diff_sec_val = df.loc[df_idx, 'time_difference_seconds']
                    
                    # Проверяем условия окрашивания
                    highlight_row = False
                    
                    # Условие 1: Разница расстояния <= -40 км
                    if dist_diff_val is not None and pd.notna(dist_diff_val):
                        try:
                            if float(dist_diff_val) <= DISTANCE_THRESHOLD:
                                highlight_row = True
                        except (ValueError, TypeError):
                            pass # Игнорируем ошибки преобразования
                            
                    # Условие 2: Разница времени <= -1ч 30мин (если первое условие не выполнено)
                    if not highlight_row and time_diff_sec_val is not None and pd.notna(time_diff_sec_val):
                         try:
                             if int(time_diff_sec_val) <= TIME_THRESHOLD_SECONDS:
                                 highlight_row = True
                         except (ValueError, TypeError):
                            pass # Игнорируем ошибки преобразования

                    # Применяем заливку ко всей строке, если условие выполнено
                    if highlight_row:
                        # print(f"   Highlighting row {row_idx}...") # Лог для отладки
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = light_red_fill
                            
                except KeyError as ke:
                     print(f"   Warning: Column {ke} not found in original DataFrame df for index {df_idx}. Skipping formatting for row {row_idx}.")
                except IndexError:
                     print(f"   Warning: Index {df_idx} out of bounds for original DataFrame df. Skipping formatting for row {row_idx}.")
            
            print(f"   Conditional formatting applied based on difference thresholds.")
            # --- КОНЕЦ УСЛОВНОГО ФОРМАТИРОВАНИЯ --- 

            # --- НАСТРОЙКА ГРАНИЦ --- 
            # Определяем стили границ
            thin_black_border = Border(left=Side(style='thin', color='FF000000'), 
                                     right=Side(style='thin', color='FF000000'), 
                                     top=Side(style='thin', color='FF000000'), 
                                     bottom=Side(style='thin', color='FF000000'))
            thin_gray_border = Border(left=Side(style='thin', color='FFA9A9A9'),  # <<< Изменен цвет на более темный серый
                                    right=Side(style='thin', color='FFA9A9A9'), 
                                    top=Side(style='thin', color='FFA9A9A9'), 
                                    bottom=Side(style='thin', color='FFA9A9A9'))

            # Применяем границы к заголовкам
            for cell in worksheet["1:1"]: # Первая строка
                 cell.border = thin_black_border
            
            # Применяем границы к ячейкам данных
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_gray_border
            print(f"   Borders applied to header and data cells.")
            # --- КОНЕЦ НАСТРОЙКИ ГРАНИЦ --- 

            # --- НАСТРОЙКА ШИРИНЫ И ПЕРЕНОСА --- 
            # workbook = writer.book # Уже получено выше
            column_widths_px = {
                'Маршрут': 380,
                'ФИО водителя': 300, # <--- ЗАДАНА ШИРИНА
                'Расстояние (км)': 140,
                'Время на маршруте': 140,
                'Расстояние (отчет)': 140,
                'Время (отчет)': 140, # <-- Добавлено
                'Разница (км)': 120,
                'Разница (время)': 120,
                'Кол-во точек': 120
            }
            
            # Применяем ширину
            for i, column_name in enumerate(export_df.columns):
                column_letter = get_column_letter(i + 1) # +1 т.к. нумерация с 1
                width_px = column_widths_px.get(column_name, 100) # Ширина по умолчанию 100px
                # Примерный перевод пикселей в ширину Excel
                excel_width = width_px / 9.0 
                worksheet.column_dimensions[column_letter].width = excel_width
                print(f"   Set width for column '{column_name}' ({column_letter}) to {excel_width:.2f}")
                
            # Включаем перенос текста для всех ячеек
            wrap_alignment = Alignment(wrap_text=True, vertical='top') # Выравнивание по верху
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = wrap_alignment
            print("   Applied text wrap to all cells.")
            # --- КОНЕЦ НАСТРОЙКИ ШИРИНЫ И ПЕРЕНОСА ---

        buffer.seek(0)
        print("   Excel buffer created.")

        # 3. Готовим имя файла и возвращаем ответ
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"summary_export_{timestamp}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        print(f"   Returning file: {filename}")
        return StreamingResponse(
            buffer,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )

    except Exception as e:
        import traceback
        print(f"!!! Error during summary export: {e}")
        print(traceback.format_exc()) # Печатаем полный traceback для диагностики
        raise HTTPException(status_code=500, detail=f"Ошибка при генерации Excel файла: {e}")

# --- КОНЕЦ НОВОГО ENDPOINT'А ---

if __name__ == "__main__":
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
